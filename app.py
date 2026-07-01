from __future__ import annotations

import hashlib
import hmac
from io import BytesIO
import logging
import os
import secrets
from datetime import datetime, timedelta, timezone
from html import escape
from pathlib import Path

from flask import Flask, jsonify, make_response, redirect, request
from google.cloud import firestore, storage

from create_report import DEFAULT_TEMPLATE, generate_report, preview_default_query_mapping
from distribution import (
    add_delivery_version,
    archive_report_definition,
    create_delivery_record,
    create_report_definition,
    download_report_definition_template,
    find_delivery_by_token,
    get_current_version,
    get_report_definition,
    list_delivery_records,
    list_download_log_records,
    list_report_definitions,
    log_download,
    make_signed_download_url,
    publish_report_definition_template,
    rollback_report_definition_template,
    update_report_definition,
    render_download_form,
    set_delivery_active,
    validate_delivery_access,
)
from mail_provider import MailDeliveryError
from mail_runtime import send_otp_pin_email

app = Flask(__name__)

TEMPLATE_PREVIEW_MAX_BYTES = int(os.environ.get("TEMPLATE_PREVIEW_MAX_BYTES", str(8 * 1024 * 1024)))
RUNTIME_TEMPLATE_DIR = os.environ.get("RUNTIME_TEMPLATE_DIR", "/tmp/ice-report-templates")


class RuntimeTemplateError(Exception):
    def __init__(self, message: str, status_code: int, reason: str) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.reason = reason


def _bigquery_project_id() -> str:
    return (
        os.environ.get("BIGQUERY_PROJECT_ID")
        or os.environ.get("PROJECT_ID")
        or os.environ.get("GOOGLE_CLOUD_PROJECT")
        or ""
    )


def _default_template_path() -> Path:
    return Path(os.environ.get("TEMPLATE_PATH", str(DEFAULT_TEMPLATE)))


def _resolve_generation_template(payload: dict) -> dict:
    report_id = str(payload.get("report_id") or "").strip()
    if not report_id:
        return {
            "template_path": _default_template_path(),
            "runtime_template": None,
            "local_path": None,
        }

    try:
        resolved = download_report_definition_template(
            report_id,
            destination_dir=RUNTIME_TEMPLATE_DIR,
        )
    except ValueError as exc:
        reason = str(exc) or "template_resolve_failed"
        logging.warning(
            "ICE_REPORT_TEMPLATE_RESOLVE_FAILED report_id=%s reason=%s",
            report_id,
            reason,
        )
        raise RuntimeTemplateError(reason, 400, reason) from exc
    except Exception as exc:
        logging.error(
            "ICE_REPORT_TEMPLATE_DOWNLOAD_FAILED report_id=%s",
            report_id,
        )
        raise RuntimeTemplateError(
            "published template could not be loaded",
            500,
            "template_download_failed",
        ) from exc

    return {
        "template_path": Path(resolved["local_path"]),
        "runtime_template": resolved.get("template"),
        "local_path": resolved.get("local_path"),
    }


def _cleanup_runtime_template(local_path: str | None) -> None:
    if not local_path:
        return
    try:
        Path(local_path).unlink(missing_ok=True)
    except Exception:
        logging.warning("ICE_REPORT_TEMPLATE_CLEANUP_FAILED")


def _env_flag(name: str) -> bool:
    return os.environ.get(name, "").strip().lower() in ("1", "true", "yes", "y", "on")


def _is_cloud_run_runtime() -> bool:
    return bool(
        os.environ.get("K_SERVICE")
        or os.environ.get("K_REVISION")
        or os.environ.get("K_CONFIGURATION")
    )


def _admin_auth_fail_closed() -> bool:
    return _is_cloud_run_runtime() or _env_flag("ADMIN_AUTH_FAIL_CLOSED")


def _log_admin_auth_failure(reason: str) -> None:
    detail = {
        "path": request.path,
        "method": request.method,
        "has_admin_key_header": bool(request.headers.get("X-Admin-Key")),
        "has_iap_user_email_header": bool(request.headers.get("X-Goog-Authenticated-User-Email")),
        "iap_auth_enabled": _admin_iap_auth_enabled(),
        "cloud_run": _is_cloud_run_runtime(),
    }
    _log_security_event(
        event_type="admin_auth_failed",
        reason=reason,
        detail=detail,
    )
    _log_admin_audit_event(
        action="admin_auth",
        result="failure",
        target_type="admin_api",
        status_code=401,
        reason=reason,
        detail=detail,
    )


@app.route("/healthz", methods=["GET"], strict_slashes=False)
@app.route("/healthz/", methods=["GET"], strict_slashes=False)
def healthz():
    return jsonify({"status": "ok"})


@app.route("/api-health", methods=["GET"], strict_slashes=False)
@app.route("/api-health/", methods=["GET"], strict_slashes=False)
def api_health():
    return jsonify({"status": "ok"})


def _check_admin() -> tuple[bool, tuple | None]:
    iap_auth = _check_admin_iap()
    if iap_auth:
        return True, None

    expected = os.environ.get("ADMIN_API_KEY")
    if not expected:
        if _admin_auth_fail_closed():
            if _admin_iap_auth_enabled():
                reason = _admin_iap_auth_failure_reason()
                logging.warning(
                    "ICE_REPORT_ADMIN_IAP_AUTH_FAILED path=%s method=%s reason=%s",
                    request.path,
                    request.method,
                    reason,
                )
                _log_admin_auth_failure(reason)
                return False, (jsonify({"error": "unauthorized"}), 401)

            logging.error(
                "ICE_REPORT_ADMIN_AUTH_NOT_CONFIGURED path=%s method=%s",
                request.path,
                request.method,
            )
            _log_admin_auth_failure("admin_key_not_configured")
            return False, (jsonify({"error": "unauthorized"}), 401)
        return True, None

    provided = request.headers.get("X-Admin-Key")
    if provided and hmac.compare_digest(provided, expected):
        return True, None

    _log_admin_auth_failure("missing_admin_key_header" if not provided else "invalid_admin_key")
    return False, (jsonify({"error": "unauthorized"}), 401)


def _admin_iap_allowed_emails() -> set[str]:
    raw = os.environ.get("ADMIN_IAP_ALLOWED_EMAILS", "")
    return {
        _normalize_email(item)
        for item in raw.replace(";", ",").split(",")
        if _normalize_email(item)
    }


def _admin_iap_auth_enabled() -> bool:
    if not _env_flag("ADMIN_IAP_AUTH_ENABLED"):
        return False

    if not _admin_iap_allowed_emails():
        return False

    current_service = os.environ.get("K_SERVICE", "").strip()
    expected_service = os.environ.get("ADMIN_IAP_SERVICE_NAME", "").strip() or "report-generator-admin"
    if not current_service:
        return False

    return current_service == expected_service


def _request_iap_email() -> str:
    raw = request.headers.get("X-Goog-Authenticated-User-Email", "")
    if ":" in raw:
        raw = raw.split(":", 1)[1]
    return _normalize_email(raw)


def _check_admin_iap() -> bool:
    if not _admin_iap_auth_enabled():
        return False

    email = _request_iap_email()
    return bool(email and email in _admin_iap_allowed_emails())


def _admin_iap_auth_failure_reason() -> str:
    if not request.headers.get("X-Goog-Authenticated-User-Email"):
        return "iap_user_email_header_missing"

    email = _request_iap_email()
    if not email:
        return "iap_user_email_invalid"

    return "iap_user_email_not_allowed"


def render_admin_ui() -> str:
    return r"""
<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>ICEレポート配布管理</title>
  <style>
    :root {
      color-scheme: light dark;
      --bg: #f5f7fb;
      --bg-grad-1: rgba(36, 87, 214, 0.10);
      --bg-grad-2: rgba(21, 115, 71, 0.08);
      --panel: rgba(255, 255, 255, 0.92);
      --panel-solid: #ffffff;
      --panel-soft: #f8fafc;
      --text: #172033;
      --muted: #667085;
      --line: #d8dee9;
      --primary: #2457d6;
      --primary-dark: #1c45ab;
      --primary-soft: rgba(36, 87, 214, 0.10);
      --danger: #c73535;
      --danger-dark: #a92b2b;
      --danger-bg: #fff1f1;
      --success: #157347;
      --success-bg: #eaf7ef;
      --warning: #966600;
      --warning-bg: #fff7df;
      --code-bg: #f2f4f7;
      --code-text: #24324a;
      --pre-bg: #101828;
      --pre-text: #e6edf7;
      --row-hover: #fbfcff;
      --focus: rgba(36, 87, 214, 0.20);
      --shadow: 0 18px 40px rgba(20, 32, 55, 0.10);
      --radius: 16px;
    }

    @media (prefers-color-scheme: dark) {
      :root {
        --bg: #0b1020;
        --bg-grad-1: rgba(91, 141, 255, 0.20);
        --bg-grad-2: rgba(35, 197, 130, 0.12);
        --panel: rgba(18, 25, 43, 0.88);
        --panel-solid: #12192b;
        --panel-soft: #182238;
        --text: #e6edf7;
        --muted: #9aa8bd;
        --line: #2d3a53;
        --primary: #7aa2ff;
        --primary-dark: #5f8df0;
        --primary-soft: rgba(122, 162, 255, 0.14);
        --danger: #ff7b7b;
        --danger-dark: #ff6262;
        --danger-bg: rgba(255, 123, 123, 0.13);
        --success: #65d99a;
        --success-bg: rgba(101, 217, 154, 0.13);
        --warning: #f6c85f;
        --warning-bg: rgba(246, 200, 95, 0.15);
        --code-bg: #0f172a;
        --code-text: #dbe7ff;
        --pre-bg: #050816;
        --pre-text: #dbe7ff;
        --row-hover: rgba(122, 162, 255, 0.08);
        --focus: rgba(122, 162, 255, 0.28);
        --shadow: 0 20px 48px rgba(0, 0, 0, 0.35);
      }
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      background:
        radial-gradient(circle at top left, var(--bg-grad-1), transparent 34rem),
        radial-gradient(circle at top right, var(--bg-grad-2), transparent 28rem),
        var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.55;
    }

    header {
      position: sticky;
      top: 0;
      z-index: 20;
      background: color-mix(in srgb, var(--bg) 88%, transparent);
      backdrop-filter: blur(14px);
      border-bottom: 1px solid var(--line);
    }

    .header-inner {
      max-width: 1280px;
      margin: 0 auto;
      padding: 18px 24px;
      display: flex;
      justify-content: space-between;
      gap: 16px;
      align-items: center;
    }

    h1 { margin: 0; font-size: 24px; letter-spacing: -0.02em; }
    h2 { margin: 0 0 16px; font-size: 18px; }
    h3 { margin: 0 0 12px; font-size: 15px; }

    .sub { margin: 4px 0 0; color: var(--muted); font-size: 13px; }

    main {
      max-width: 1280px;
      margin: 0 auto;
      padding: 24px;
    }

    .grid {
      display: grid;
      grid-template-columns: minmax(320px, 420px) minmax(0, 1fr);
      gap: 20px;
      align-items: start;
    }

    .card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 18px;
      margin-bottom: 20px;
      backdrop-filter: blur(10px);
    }

    .toolbar {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
      margin-bottom: 14px;
    }

    .field { margin-bottom: 12px; }
    label { display: block; font-size: 12px; color: var(--muted); margin-bottom: 5px; }
    .help { margin: -6px 0 12px; color: var(--muted); font-size: 12px; }
    .muted { color: var(--muted); font-size: 13px; }

    input, select {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px 11px;
      font: inherit;
      background: var(--panel-solid);
      color: var(--text);
    }

    input::placeholder { color: color-mix(in srgb, var(--muted) 75%, transparent); }

    input:focus, select:focus {
      outline: 3px solid var(--focus);
      border-color: var(--primary);
    }

    button {
      border: 1px solid transparent;
      border-radius: 10px;
      padding: 9px 12px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      color: #fff;
      background: var(--primary);
      transition: 0.12s ease;
    }

    button:hover { background: var(--primary-dark); transform: translateY(-1px); }
    button:disabled { opacity: 0.55; cursor: not-allowed; transform: none; }

    button.secondary {
      color: var(--text);
      background: var(--panel-solid);
      border-color: var(--line);
    }

    button.secondary:hover { background: var(--panel-soft); }

    button.danger { background: var(--danger); color: #fff; }
    button.danger:hover { background: var(--danger-dark); }

    button.small {
      padding: 6px 9px;
      font-size: 12px;
      border-radius: 8px;
    }

    .row-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 7px;
    }

    .inline-fields {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
    }

    .mono { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; }

    .status-pill {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 3px 9px;
      font-size: 12px;
      font-weight: 800;
      border: 1px solid transparent;
    }

    .status-active {
      color: var(--success);
      background: var(--success-bg);
      border-color: color-mix(in srgb, var(--success) 35%, transparent);
    }

    .status-disabled {
      color: var(--danger);
      background: var(--danger-bg);
      border-color: color-mix(in srgb, var(--danger) 35%, transparent);
    }

    .status-warning {
      color: var(--warning);
      background: var(--warning-bg);
      border-color: color-mix(in srgb, var(--warning) 35%, transparent);
    }

    .table-wrap {
      width: 100%;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: var(--panel-solid);
    }

    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }

    th, td {
      border-bottom: 1px solid var(--line);
      padding: 10px;
      text-align: left;
      vertical-align: top;
    }

    th {
      background: var(--panel-soft);
      position: sticky;
      top: 0;
      z-index: 1;
      color: var(--muted);
      font-size: 12px;
      white-space: nowrap;
    }

    tr:hover td { background: var(--row-hover); }
    tr:last-child td { border-bottom: 0; }

    a { color: var(--primary); }

    code {
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      word-break: break-all;
      color: var(--code-text);
      background: var(--code-bg);
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 2px 5px;
    }

    pre {
      white-space: pre-wrap;
      word-break: break-word;
      background: var(--pre-bg);
      color: var(--pre-text);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 12px;
      min-height: 44px;
      font-size: 12px;
    }

    .notice {
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 12px;
      background: var(--panel-soft);
      color: var(--muted);
      font-size: 13px;
    }

    .summary-cards {
      display: grid;
      grid-template-columns: repeat(4, minmax(120px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }

    .summary-card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 12px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(10px);
    }

    .summary-card .label { color: var(--muted); font-size: 12px; }
    .summary-card .value { font-size: 22px; font-weight: 850; margin-top: 4px; }

    .version-panel {
      background: var(--primary-soft);
      border: 1px dashed color-mix(in srgb, var(--primary) 45%, transparent);
      border-radius: 12px;
      padding: 12px;
    }

    .delivery-table {
      min-width: 1080px;
      table-layout: fixed;
    }

    .delivery-id-cell {
      width: 110px;
    }

    .delivery-id-cell code {
      display: block;
      overflow-wrap: anywhere;
    }

    .delivery-meta-cell {
      width: 140px;
    }

    .delivery-status-cell {
      width: 88px;
    }

    .delivery-url-cell {
      width: 280px;
    }

    .delivery-url-link {
      display: block;
      overflow-wrap: anywhere;
      word-break: break-word;
    }

    .delivery-version-cell {
      width: 360px;
    }

    .delivery-actions-cell {
      width: 160px;
    }

    .definition-table {
      min-width: 980px;
      table-layout: fixed;
    }

    .definition-id-cell {
      width: 150px;
    }

    .definition-name-cell {
      width: 190px;
    }

    .definition-storage-cell {
      width: 260px;
    }

    .delivery-uri-toggle {
      margin-top: 6px;
      font-size: 12px;
      color: var(--muted);
    }

    .delivery-uri-toggle summary {
      cursor: pointer;
    }

    .delivery-uri-toggle code {
      display: block;
      margin-top: 6px;
      white-space: pre-wrap;
    }

    .toast {
      position: fixed;
      right: 20px;
      bottom: 20px;
      background: var(--pre-bg);
      color: var(--pre-text);
      border: 1px solid var(--line);
      padding: 10px 14px;
      border-radius: 10px;
      box-shadow: var(--shadow);
      opacity: 0;
      transform: translateY(8px);
      pointer-events: none;
      transition: 0.16s ease;
      z-index: 100;
    }

    .toast.show { opacity: 1; transform: translateY(0); }

    details.faq {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 0;
      margin-bottom: 20px;
      overflow: hidden;
      backdrop-filter: blur(10px);
    }

    details.faq > summary {
      cursor: pointer;
      list-style: none;
      padding: 16px 18px;
      font-size: 18px;
      font-weight: 850;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      border-bottom: 1px solid transparent;
    }

    details.faq > summary::-webkit-details-marker { display: none; }

    details.faq > summary::after {
      content: "開く";
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 3px 9px;
      background: var(--panel-solid);
    }

    details.faq[open] > summary { border-bottom-color: var(--line); }
    details.faq[open] > summary::after { content: "閉じる"; }

    .faq-body { padding: 16px 18px 18px; }

    .faq-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }

    .faq-item {
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 12px;
      background: var(--panel-soft);
    }

    .faq-item h3 { margin-bottom: 8px; }
    .faq-item ul { margin: 8px 0 0 18px; padding: 0; }
    .faq-item li { margin: 4px 0; }

    .faq-note {
      margin-top: 14px;
      border-left: 4px solid var(--primary);
      padding: 10px 12px;
      border-radius: 10px;
      background: var(--primary-soft);
      color: var(--text);
      font-size: 13px;
    }


    @media (max-width: 980px) {
      .grid { grid-template-columns: 1fr; }
      .summary-cards { grid-template-columns: repeat(2, minmax(120px, 1fr)); }
      .faq-grid { grid-template-columns: 1fr; }
      .header-inner { align-items: flex-start; flex-direction: column; }
    }
  </style>
</head>
<body>
<header>
  <div class="header-inner">
    <div>
      <h1>ICEレポート配布管理</h1>
      <p class="sub">対象レポート: ジャンプ＋デジタルコミックス月次データ</p>
      <p class="sub">配布URL作成、version更新、停止/再開、DLログ確認</p>
    </div>
    <div class="toolbar" style="margin:0;">
      <button class="secondary" onclick="loadAll()">全体更新</button>
      <button class="secondary" onclick="clearAdminKeyAndReload()">管理キー再入力</button>
    </div>
  </div>
</header>

<main>
  <div class="summary-cards">
    <div class="summary-card"><div class="label">配布総数</div><div class="value" id="summaryTotal">-</div></div>
    <div class="summary-card"><div class="label">active</div><div class="value" id="summaryActive">-</div></div>
    <div class="summary-card"><div class="label">disabled</div><div class="value" id="summaryDisabled">-</div></div>
    <div class="summary-card"><div class="label">定義 / 表示ログ</div><div class="value" id="summaryDefinitionsLogs">- / -</div></div>
  </div>



  <details class="faq">
    <summary>使い方・FAQ</summary>
    <div class="faq-body">
      <div class="faq-grid">
        <div class="faq-item">
          <h3>配布URLを作成</h3>
          <p class="muted">新しい配布レコードを作成します。GCS URIが空欄の場合は、BigQueryを再実行してExcelを生成し、GCSへ保存してから配布URLを発行します。</p>
          <ul class="muted">
            <li>顧客名が空欄: <code>customer_name is required</code> エラー。</li>
            <li>対象月が空欄: <code>report_month is required</code> エラー。</li>
            <li>許可ドメインが空欄: 既定ドメインを自動適用。</li>
            <li>生成ファイル名が空欄: 当日日付を使った標準ファイル名で生成。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>入力フォーマット</h3>
          <ul class="muted">
            <li>対象月: 運用上は <code>YYYY-MM</code> 形式で入力します。</li>
            <li>許可メール: <code>user@example.com, user2@example.com</code> のようにカンマ区切り。</li>
            <li>許可ドメイン: <code>example.co.jp, example.com</code> のようにカンマ区切り。</li>
            <li>GCS URI: <code>gs://bucket/path/file.xlsx</code> 形式。形式外は backend 側でエラーになります。</li>
            <li>生成ファイル名: 指定する場合は <code>.xlsx</code> 終端必須。version追加時は形式外だとエラーになります。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>最新GCSファイル一覧</h3>
          <p class="muted">指定prefix配下の最新Excelファイルを表示します。<code>配布作成に使う</code> を押すと、そのGCS URIを配布作成フォームへ反映します。</p>
          <ul class="muted">
            <li>prefixが空欄に近い場合、探索範囲が広くなります。</li>
            <li><code>.xlsx</code> 以外のファイルは一覧対象外です。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>配布一覧・検索・状態フィルタ</h3>
          <p class="muted">配布レコードを一覧表示します。検索欄では顧客名、delivery_id、対象月、ファイル名、URLを絞り込めます。</p>
          <ul class="muted">
            <li><code>active</code>: ダウンロード可能な状態。</li>
            <li><code>disabled</code>: 停止状態。ユーザーがアクセスしてもダウンロード不可。</li>
            <li>表示上の検索・フィルタは画面内データに対する絞り込みです。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>停止 / 再開</h3>
          <p class="muted">停止は配布レコードの <code>active=false</code>、再開は <code>active=true</code> に更新します。配布URL自体は変わりません。</p>
          <ul class="muted">
            <li>停止中: 許可メールでもダウンロード不可。</li>
            <li>再開後: 期限内かつ許可メール/ドメインに一致すれば再びダウンロード可能。</li>
            <li>期限切れの場合、再開しても期限判定で拒否されます。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>version追加 / overwrite</h3>
          <p class="muted">対象deliveryのクエリを再実行し、Excelを再生成して新しいversionとして追加します。配布URLは変わらず、current_versionだけが更新されます。</p>
          <ul class="muted">
            <li>overwrite OFF: 入力した保存ファイル名で新しいファイルを作成。</li>
            <li>overwrite ON: 現在versionのファイル名を再利用して上書き保存。</li>
            <li>保存ファイル名が空欄かつoverwrite OFF: エラー。</li>
            <li><code>.xlsx</code> 以外: エラー。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>ログ</h3>
          <p class="muted">ダウンロード成功時の履歴を表示します。delivery単位のログボタンを押すと、そのdelivery_idで絞り込んだログを表示します。</p>
          <ul class="muted">
            <li>記録対象: 日時、delivery_id、顧客名、対象月、email、version、file。</li>
            <li>ログ検索欄は画面内に読み込まれたログに対する絞り込みです。</li>
          </ul>
        </div>

        <div class="faq-item">
          <h3>管理キー・エラー時</h3>
          <p class="muted">初回アクセス時に管理キーを入力します。キーはブラウザのlocalStorageに保存され、API呼び出し時に <code>X-Admin-Key</code> として送信されます。</p>
          <ul class="muted">
            <li>401: 管理キー不正。保存キーを削除して再入力します。</li>
            <li>BigQuery / GCS / Firestore エラー: 画面の結果欄にbackendエラーを表示します。</li>
            <li>操作中はボタンを無効化し、二重実行を抑止します。</li>
          </ul>
        </div>
      </div>

      <div class="faq-note">
        月次運用では、配布作成前に対象月・顧客名・許可宛先・保存ファイル名を確認します。overwrite ONは現在versionのGCS objectを上書きするため、既存ファイルのGoogle Drive backupと承認記録がある場合だけ使います。期限切れ配布はcleanupでactive=falseにし、GCS削除はbackup後の別承認で扱います。
      </div>
    </div>
  </details>

  <div class="card">
    <h2>レポート定義</h2>
    <div class="inline-fields">
      <div class="field"><label>report_id</label><input id="definitionId" placeholder="例: plus-monthly-downloads"></div>
      <div class="field"><label>name</label><input id="definitionName" placeholder="レポート名"></div>
    </div>
    <div class="inline-fields">
      <div class="field"><label>owner</label><input id="definitionOwner" placeholder="システム管理室"></div>
      <div class="field"><label>primary operator</label><input id="definitionOperator" placeholder="篠原邦昭"></div>
    </div>
    <div class="inline-fields">
      <div class="field"><label>customer</label><input id="definitionCustomer" placeholder="顧客 / recipient group"></div>
      <div class="field"><label>default month</label><input id="definitionDefaultMonth" placeholder="YYYY-MM"></div>
    </div>
    <div class="inline-fields">
      <div class="field"><label>GCS prefix</label><input id="definitionGcsPrefix" placeholder="reports/plus/"></div>
      <div class="field"><label>Drive folder name</label><input id="definitionDriveFolder" placeholder="OMFダウンロード数報告"></div>
    </div>
    <div class="field"><label>initial version note</label><input id="definitionVersionNote" placeholder="initial definition"></div>
    <div class="toolbar">
      <button id="definitionCreateButton" onclick="createReportDefinition()">定義を追加</button>
      <button class="secondary" id="definitionUpdateButton" onclick="updateReportDefinition()">定義を更新</button>
      <button class="danger" id="definitionArchiveButton" onclick="archiveReportDefinition()">archive</button>
      <button class="secondary" onclick="clearDefinitionForm()">入力クリア</button>
    </div>
    <pre id="definitionResult">待機中</pre>
    <div class="field"><label>Excel template preview (.xlsx)</label><input id="templatePreviewFile" type="file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"></div>
    <div class="inline-fields">
      <div class="field"><label>template version note</label><input id="templateVersionNote" placeholder="template update note"></div>
      <div class="field"><label>rollback version</label><input id="templateRollbackVersion" placeholder="version number"></div>
    </div>
    <div class="toolbar">
      <button class="secondary" id="templatePreviewButton" onclick="previewReportTemplate()">template preview</button>
      <button class="secondary" id="templatePublishButton" onclick="publishReportTemplate()">template publish</button>
      <button class="secondary" id="templateRollbackButton" onclick="rollbackReportTemplate()">template rollback</button>
    </div>
    <pre id="templatePreviewResult">not loaded</pre>
    <div class="toolbar">
      <button class="secondary" id="queryMappingPreviewButton" onclick="previewQueryMapping()">query / mapping dry-run</button>
    </div>
    <pre id="queryMappingPreviewResult">not loaded</pre>
    <div class="toolbar">
      <input id="definitionSearch" placeholder="report_id / name / owner / GCS prefixで検索" oninput="renderDefinitionsFromState()" style="min-width:260px;flex:1;">
      <select id="definitionStatusFilter" onchange="renderDefinitionsFromState()" style="width:150px;">
        <option value="all">all</option>
        <option value="active">active</option>
        <option value="archived">archived</option>
      </select>
      <button class="secondary" onclick="loadReportDefinitions()">定義一覧を更新</button>
    </div>
    <div id="reportDefinitions" class="notice">loading...</div>
  </div>


  <div class="grid">
    <section>
      <div class="card">
        <h2>配布作成</h2>
        <p class="muted">対象レポート: ジャンプ＋デジタルコミックス月次データ</p>
        <div class="inline-fields">
          <div class="field"><label>顧客名</label><input id="createCustomer" placeholder="顧客名" value="一ツ橋企画"></div>
          <div class="field"><label>対象月</label><input id="createMonth" placeholder="YYYY-MM" value="2026-04"></div>
        </div>
        <div class="field"><label>許可メール カンマ区切り</label><input id="createEmails" placeholder="user@example.com, user2@example.com"></div>
        <div class="field"><label>許可ドメイン カンマ区切り</label><input id="createDomains" placeholder="空欄の場合は既定ドメインを使用"></div>
        <p class="help">空欄の場合は shueisha.co.jp, sur.co.jp, hitotsubashi.co.jp, impress.co.jp を許可します。</p>
        <div class="field"><label>GCS URI。空欄ならクエリ再実行して生成</label><input id="createGcs" placeholder="gs://ice-report-files/reports/plus/...xlsx"></div>
        <div class="field"><label>生成ファイル名 .xlsx</label><input id="createOutputFilename" placeholder="空欄の場合は当日日付で自動生成"></div>
        <p class="help">生成ファイル名を空欄にすると、実行当日の日付を使った標準ファイル名で生成します。</p>
        <div class="toolbar">
          <button id="createDeliveryButton" onclick="createDelivery()">配布URLを作成</button>
          <button class="secondary" onclick="clearCreateForm()">入力クリア</button>
        </div>
        <pre id="createResult">待機中</pre>
      </div>

      <div class="card">
        <h2>最新GCSファイル一覧</h2>
        <div class="field"><label>prefix</label><input id="gcsPrefix" value="reports/plus/"></div>
        <div class="toolbar">
          <button onclick="loadGcsFiles()">最新ファイルを表示</button>
        </div>
        <div id="gcsFiles" class="notice">未読み込み</div>
      </div>
    </section>

    <section>
      <div class="card">
        <h2>配布一覧</h2>
        <div class="toolbar">
          <input id="deliverySearch" placeholder="顧客名 / delivery_id / 月 / ファイル名 / URLで検索" oninput="renderDeliveriesFromState()" style="min-width:260px;flex:1;">
          <select id="deliveryStatusFilter" onchange="renderDeliveriesFromState()" style="width:150px;">
            <option value="all">all</option>
            <option value="active">active</option>
            <option value="disabled">disabled</option>
          </select>
          <button class="secondary" onclick="loadDeliveries()">一覧を更新</button>
        </div>
        <div id="deliveries" class="notice">loading...</div>
      </div>

      <div class="card">
        <h2>ダウンロードログ</h2>
        <div class="toolbar">
          <input id="logSearch" placeholder="delivery_id / email / 顧客名 / fileで絞り込み" oninput="renderLogsFromState()" style="min-width:260px;flex:1;">
          <button class="secondary" onclick="loadLogs()">全ログを更新</button>
        </div>
        <div id="logs" class="notice">loading...</div>
      </div>
    </section>
  </div>
</main>

<div id="toast" class="toast"></div>

<script>
const baseUrl = window.location.origin;
const DEFAULT_ALLOWED_DOMAINS = ["shueisha.co.jp", "sur.co.jp", "hitotsubashi.co.jp", "impress.co.jp"];
let createDeliveryInProgress = false;
let reportDefinitionInProgress = false;
let templatePreviewInProgress = false;
const versionInProgress = {};
const ADMIN_KEY_STORAGE = "ice_admin_api_key";
let reportDefinitionItems = [];
const reportDefinitionDetails = {};
let deliveryItems = [];
let logItems = [];

function getAdminKey() {
  let key = localStorage.getItem(ADMIN_KEY_STORAGE);
  if (!key) {
    key = prompt("管理用APIキーを入力してください");
    if (key) {
      localStorage.setItem(ADMIN_KEY_STORAGE, key);
    }
  }
  return key || "";
}

function clearAdminKeyAndReload() {
  localStorage.removeItem(ADMIN_KEY_STORAGE);
  location.reload();
}

function splitList(value) {
  return (value || "")
    .split(",")
    .map(item => item.trim())
    .filter(Boolean);
}

async function api(path, options = {}) {
  const headers = Object.assign({}, options.headers || {});
  const adminKey = getAdminKey();
  if (adminKey) {
    headers["X-Admin-Key"] = adminKey;
  }
  if (options.body && !(options.body instanceof FormData) && !headers["Content-Type"]) {
    headers["Content-Type"] = "application/json";
  }

  const res = await fetch(baseUrl + path, {
    method: options.method || "GET",
    headers,
    body: options.body,
  });

  const text = await res.text();
  let data = {};

  try {
    data = text ? JSON.parse(text) : {};
  } catch (e) {
    data = {raw: text};
  }

  if (!res.ok) {
    const message = data && (data.error || data.message || data.raw) ? (data.error || data.message || data.raw) : ("HTTP " + res.status);
    if (res.status === 401) {
      localStorage.removeItem(ADMIN_KEY_STORAGE);
    }
    throw new Error(message);
  }

  return data;
}

function esc(s) {
  return String(s ?? "")
    .replace(/&/g, "&amp;")
    .replace(/</g, "&lt;")
    .replace(/>/g, "&gt;")
    .replace(/"/g, "&quot;")
    .replace(/'/g, "&#39;");
}

function attr(s) {
  return esc(s).replace(/`/g, "&#96;");
}

function formatSize(size) {
  const n = Number(size || 0);
  if (n >= 1024 * 1024) {
    return (n / 1024 / 1024).toFixed(1) + " MB";
  }
  if (n >= 1024) {
    return (n / 1024).toFixed(1) + " KB";
  }
  return n + " B";
}

function formatDateTime(value) {
  if (!value) {
    return "-";
  }
  const d = new Date(value);
  if (Number.isNaN(d.getTime())) {
    return value;
  }
  return d.toLocaleString("ja-JP", {hour12: false});
}

function showToast(message) {
  const el = document.getElementById("toast");
  el.textContent = message;
  el.classList.add("show");
  setTimeout(() => el.classList.remove("show"), 1600);
}

function copyText(text) {
  navigator.clipboard.writeText(text || "").then(() => showToast("コピーしました")).catch(() => showToast("コピーに失敗しました"));
}

function clearCreateForm() {
  document.getElementById("createCustomer").value = "";
  document.getElementById("createMonth").value = "";
  document.getElementById("createEmails").value = "";
  document.getElementById("createDomains").value = "";
  document.getElementById("createGcs").value = "";
  document.getElementById("createOutputFilename").value = "";
  document.getElementById("createResult").textContent = "待機中";
}

async function createDelivery() {
  if (createDeliveryInProgress) {
    return;
  }

  createDeliveryInProgress = true;

  const button = document.getElementById("createDeliveryButton");
  const resultEl = document.getElementById("createResult");

  button.disabled = true;
  button.textContent = "作成中...";
  resultEl.textContent = "配布URLを作成中です。クエリ実行・Excel生成・GCS保存中...";

  const inputDomains = splitList(document.getElementById("createDomains").value);

  const payload = {
    customer_name: document.getElementById("createCustomer").value,
    report_month: document.getElementById("createMonth").value,
    gcs_uri: document.getElementById("createGcs").value,
    output_filename: document.getElementById("createOutputFilename").value,
    report_id: document.getElementById("definitionId").value,
    allowed_emails: splitList(document.getElementById("createEmails").value),
    allowed_domains: inputDomains.length ? inputDomains : DEFAULT_ALLOWED_DOMAINS
  };

  try {
    const data = await api("/deliveries", {
      method: "POST",
      body: JSON.stringify(payload)
    });

    const url = data.download_url && data.download_url.startsWith("http")
      ? data.download_url
      : baseUrl + (data.download_url || "");

    resultEl.textContent = "配布URLが作成されました\nURL: " + url;
    showToast("配布URLを作成しました");

    await loadDeliveries();
    await loadGcsFiles();

  } catch (e) {
    resultEl.textContent = "配布URL作成に失敗しました\n" + e.message;
  } finally {
    createDeliveryInProgress = false;
    button.disabled = false;
    button.textContent = "配布URLを作成";
  }
}

async function loadGcsFiles(targetInputId = "") {
  const el = document.getElementById("gcsFiles");
  const prefix = document.getElementById("gcsPrefix").value || "reports/plus/";

  el.innerHTML = "<p class='muted'>loading GCS files...</p>";

  try {
    const data = await api("/gcs-files?prefix=" + encodeURIComponent(prefix) + "&limit=50");
    renderGcsFiles(data.items || [], targetInputId);
  } catch (e) {
    el.innerHTML = "<p style='color:#c73535'>" + esc(e.message) + "</p>";
  }
}

function renderGcsFiles(items, targetInputId = "") {
  const el = document.getElementById("gcsFiles");

  if (!items.length) {
    el.innerHTML = "<p class='muted'>該当ファイルなし</p>";
    return;
  }

  const rows = items.map(item => {
    const useButton = targetInputId
      ? "<button class='small' onclick=\"useGcsUri('" + attr(targetInputId) + "', '" + attr(item.gcs_uri) + "')\">このURIを使う</button>"
      : "<button class='small' onclick=\"copyGcsUriToCreate('" + attr(item.gcs_uri) + "')\">配布作成に使う</button>";

    return "<tr>" +
      "<td>" + esc(formatDateTime(item.updated || "")) + "</td>" +
      "<td><code>" + esc(item.gcs_uri || "") + "</code></td>" +
      "<td>" + esc(formatSize(item.size)) + "</td>" +
      "<td><div class='row-actions'>" + useButton +
        "<button class='small secondary' onclick=\"copyText('" + attr(item.gcs_uri || "") + "')\">コピー</button>" +
      "</div></td>" +
    "</tr>";
  }).join("");

  el.innerHTML =
    "<p class='muted'>latest files: " + items.length + "件</p>" +
    "<div class='table-wrap'><table>" +
    "<thead><tr><th>更新日時</th><th>GCS URI</th><th>size</th><th>操作</th></tr></thead>" +
    "<tbody>" + rows + "</tbody></table></div>";
}

function copyGcsUriToCreate(gcsUri) {
  document.getElementById("createGcs").value = gcsUri;
  showToast("配布作成フォームに反映しました");
}

function useGcsUri(inputId, gcsUri) {
  const input = document.getElementById(inputId);
  if (input) {
    input.value = gcsUri;
    showToast("GCS URIを反映しました");
  }
}

function definitionPayload() {
  return {
    report_id: document.getElementById("definitionId").value,
    name: document.getElementById("definitionName").value,
    owner: document.getElementById("definitionOwner").value,
    primary_operator: document.getElementById("definitionOperator").value,
    customer_name: document.getElementById("definitionCustomer").value,
    default_report_month: document.getElementById("definitionDefaultMonth").value,
    gcs_prefix: document.getElementById("definitionGcsPrefix").value,
    drive_folder_name: document.getElementById("definitionDriveFolder").value,
    version_note: document.getElementById("definitionVersionNote").value
  };
}

function setDefinitionButtons(disabled) {
  [
    "definitionCreateButton",
    "definitionUpdateButton",
    "definitionArchiveButton",
    "templatePreviewButton",
    "templatePublishButton",
    "templateRollbackButton",
    "queryMappingPreviewButton"
  ].forEach(id => {
    const button = document.getElementById(id);
    if (button) {
      button.disabled = disabled;
    }
  });
}

function clearDefinitionForm() {
  [
    "definitionId",
    "definitionName",
    "definitionOwner",
    "definitionOperator",
    "definitionCustomer",
    "definitionDefaultMonth",
    "definitionGcsPrefix",
    "definitionDriveFolder",
    "definitionVersionNote",
    "templateVersionNote",
    "templateRollbackVersion"
  ].forEach(id => {
    const input = document.getElementById(id);
    if (input) {
      input.value = "";
    }
  });
  document.getElementById("definitionResult").textContent = "待機中";
}

function fillDefinitionForm(reportId) {
  const item = reportDefinitionItems.find(v => v.report_id === reportId) || reportDefinitionDetails[reportId];
  if (!item) {
    return;
  }

  document.getElementById("definitionId").value = item.report_id || "";
  document.getElementById("definitionName").value = item.name || "";
  document.getElementById("definitionOwner").value = item.owner || "";
  document.getElementById("definitionOperator").value = item.primary_operator || "";
  document.getElementById("definitionCustomer").value = item.customer_name || "";
  document.getElementById("definitionDefaultMonth").value = item.default_report_month || "";
  document.getElementById("definitionGcsPrefix").value = item.gcs_prefix || "";
  document.getElementById("definitionDriveFolder").value = item.drive_folder_name || "";
  document.getElementById("definitionVersionNote").value = "";
  document.getElementById("definitionResult").textContent = "編集中: " + (item.report_id || "");
}

async function createReportDefinition() {
  if (reportDefinitionInProgress) {
    return;
  }

  reportDefinitionInProgress = true;
  setDefinitionButtons(true);
  const resultEl = document.getElementById("definitionResult");
  resultEl.textContent = "定義を追加中...";

  try {
    const data = await api("/report-definitions", {
      method: "POST",
      body: JSON.stringify(definitionPayload())
    });
    resultEl.textContent = "定義を追加しました\n" + JSON.stringify(data.result || data.item || data, null, 2);
    showToast("定義を追加しました");
    await loadReportDefinitions();
  } catch (e) {
    resultEl.textContent = "定義追加に失敗しました\n" + e.message;
  } finally {
    reportDefinitionInProgress = false;
    setDefinitionButtons(false);
  }
}

async function updateReportDefinition() {
  if (reportDefinitionInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  reportDefinitionInProgress = true;
  setDefinitionButtons(true);
  const resultEl = document.getElementById("definitionResult");
  resultEl.textContent = "定義を更新中...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId), {
      method: "PATCH",
      body: JSON.stringify(definitionPayload())
    });
    resultEl.textContent = "定義を更新しました\n" + JSON.stringify(data.result || data.item || data, null, 2);
    showToast("定義を更新しました");
    delete reportDefinitionDetails[reportId];
    await loadReportDefinitions();
  } catch (e) {
    resultEl.textContent = "定義更新に失敗しました\n" + e.message;
  } finally {
    reportDefinitionInProgress = false;
    setDefinitionButtons(false);
  }
}

async function archiveReportDefinition() {
  if (reportDefinitionInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  if (!reportId || !confirm("archiveします: " + reportId)) {
    return;
  }

  reportDefinitionInProgress = true;
  setDefinitionButtons(true);
  const resultEl = document.getElementById("definitionResult");
  resultEl.textContent = "archive中...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId) + "/archive", {
      method: "POST"
    });
    resultEl.textContent = "archiveしました\n" + JSON.stringify(data.result || data.item || data, null, 2);
    showToast("archiveしました");
    delete reportDefinitionDetails[reportId];
    await loadReportDefinitions();
  } catch (e) {
    resultEl.textContent = "archiveに失敗しました\n" + e.message;
  } finally {
    reportDefinitionInProgress = false;
    setDefinitionButtons(false);
  }
}

async function previewReportTemplate() {
  if (templatePreviewInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  const fileInput = document.getElementById("templatePreviewFile");
  const resultEl = document.getElementById("templatePreviewResult");
  const button = document.getElementById("templatePreviewButton");

  if (!reportId) {
    resultEl.textContent = "report_id is required";
    return;
  }
  if (!fileInput.files || !fileInput.files.length) {
    resultEl.textContent = "template .xlsx file is required";
    return;
  }

  const form = new FormData();
  form.append("template_file", fileInput.files[0]);
  templatePreviewInProgress = true;
  button.disabled = true;
  resultEl.textContent = "previewing template...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId) + "/template-preview", {
      method: "POST",
      body: form
    });
    resultEl.textContent = JSON.stringify(data.preview || data, null, 2);
    showToast("template preview completed");
  } catch (e) {
    resultEl.textContent = "template preview failed\n" + e.message;
  } finally {
    templatePreviewInProgress = false;
    button.disabled = false;
  }
}

async function publishReportTemplate() {
  if (templatePreviewInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  const fileInput = document.getElementById("templatePreviewFile");
  const resultEl = document.getElementById("templatePreviewResult");
  const button = document.getElementById("templatePublishButton");

  if (!reportId) {
    resultEl.textContent = "report_id is required";
    return;
  }
  if (!fileInput.files || !fileInput.files.length) {
    resultEl.textContent = "template .xlsx file is required";
    return;
  }
  if (!confirm("publish template for " + reportId + "?")) {
    return;
  }

  const form = new FormData();
  form.append("template_file", fileInput.files[0]);
  form.append("note", document.getElementById("templateVersionNote").value || "");
  templatePreviewInProgress = true;
  button.disabled = true;
  resultEl.textContent = "publishing template...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId) + "/template-publish", {
      method: "POST",
      body: form
    });
    resultEl.textContent = "template published\n" + JSON.stringify(data.template || data.result || data, null, 2);
    showToast("template published");
    delete reportDefinitionDetails[reportId];
    await loadReportDefinitions();
  } catch (e) {
    resultEl.textContent = "template publish failed\n" + e.message;
  } finally {
    templatePreviewInProgress = false;
    button.disabled = false;
  }
}

async function rollbackReportTemplate() {
  if (templatePreviewInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  const version = document.getElementById("templateRollbackVersion").value;
  const resultEl = document.getElementById("templatePreviewResult");
  const button = document.getElementById("templateRollbackButton");

  if (!reportId) {
    resultEl.textContent = "report_id is required";
    return;
  }
  if (!version) {
    resultEl.textContent = "rollback version is required";
    return;
  }
  if (!confirm("rollback template current_version to v" + version + " for " + reportId + "?")) {
    return;
  }

  templatePreviewInProgress = true;
  button.disabled = true;
  resultEl.textContent = "rolling back template...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId) + "/template-rollback", {
      method: "POST",
      body: JSON.stringify({version: Number(version)})
    });
    resultEl.textContent = "template rolled back\n" + JSON.stringify(data.result || data.item || data, null, 2);
    showToast("template rolled back");
    delete reportDefinitionDetails[reportId];
    await loadReportDefinitions();
  } catch (e) {
    resultEl.textContent = "template rollback failed\n" + e.message;
  } finally {
    templatePreviewInProgress = false;
    button.disabled = false;
  }
}

async function previewQueryMapping() {
  if (templatePreviewInProgress) {
    return;
  }

  const reportId = document.getElementById("definitionId").value;
  const resultEl = document.getElementById("queryMappingPreviewResult");
  const button = document.getElementById("queryMappingPreviewButton");

  if (!reportId) {
    resultEl.textContent = "report_id is required";
    return;
  }

  templatePreviewInProgress = true;
  button.disabled = true;
  resultEl.textContent = "running query dry-run...";

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(reportId) + "/query-mapping-preview", {
      method: "POST",
      body: JSON.stringify({})
    });
    resultEl.textContent = JSON.stringify(data.preview || data.result || data, null, 2);
    showToast("query / mapping preview completed");
  } catch (e) {
    resultEl.textContent = "query / mapping preview failed\n" + e.message;
  } finally {
    templatePreviewInProgress = false;
    button.disabled = false;
  }
}

function definitionStatus(item) {
  return item.status || "unknown";
}

function definitionSearchText(item) {
  return [
    item.report_id,
    item.name,
    item.status,
    item.owner,
    item.primary_operator,
    item.customer_name,
    item.default_report_month,
    item.gcs_prefix,
    item.drive_folder_name,
    item.current_version
  ].join(" ").toLowerCase();
}

async function loadReportDefinitions() {
  const el = document.getElementById("reportDefinitions");
  el.innerHTML = "<p class='muted'>loading report definitions...</p>";

  try {
    const data = await api("/report-definitions?limit=100");
    reportDefinitionItems = data.items || [];
    renderDefinitionsFromState();
  } catch (e) {
    el.innerHTML = "<p style='color:#c73535'>" + esc(e.message) + "</p>";
  }
}

async function loadReportDefinitionDetail(reportId) {
  const safeId = String(reportId || "");
  const outputId = "definitionVersions_" + safeId.replace(/[^a-zA-Z0-9_-]/g, "_");
  const el = document.getElementById(outputId);

  if (el) {
    el.innerHTML = "<p class='muted'>loading versions...</p>";
  }

  try {
    const data = await api("/report-definitions/" + encodeURIComponent(safeId));
    reportDefinitionDetails[safeId] = data.item || data.result || data;
    renderDefinitionsFromState();
  } catch (e) {
    if (el) {
      el.innerHTML = "<p style='color:#c73535'>" + esc(e.message) + "</p>";
    }
  }
}

function renderDefinitionsFromState() {
  const q = (document.getElementById("definitionSearch").value || "").toLowerCase().trim();
  const status = document.getElementById("definitionStatusFilter").value;

  const filtered = reportDefinitionItems.filter(item => {
    if (status !== "all" && definitionStatus(item) !== status) {
      return false;
    }
    if (q && !definitionSearchText(item).includes(q)) {
      return false;
    }
    return true;
  });

  renderReportDefinitions(filtered);
  updateSummary();
}

function renderDefinitionVersions(item) {
  const reportId = String(item.report_id || "");
  const detail = reportDefinitionDetails[reportId];
  const outputId = "definitionVersions_" + reportId.replace(/[^a-zA-Z0-9_-]/g, "_");

  if (!detail) {
    return "<div class='row-actions' style='margin-top:10px;'>" +
      "<button class='small secondary' onclick=\"loadReportDefinitionDetail('" + attr(reportId) + "')\">versionを表示</button>" +
      "</div><div id='" + attr(outputId) + "'></div>";
  }

  const versions = detail.versions || [];
  if (!versions.length) {
    return "<div class='muted' id='" + attr(outputId) + "'>versionなし</div>";
  }

  const versionRows = versions.map(v => {
    const statusClass = v.current ? "status-active" : "status-warning";
    return "<div class='version-panel' style='margin-top:8px;'>" +
      "<div><strong>v" + esc(v.version || "-") + "</strong> <span class='status-pill " + statusClass + "'>" + esc(v.current ? "current" : (v.status || "draft")) + "</span></div>" +
      "<div class='muted'>" + esc(formatDateTime(v.updated_at || v.created_at || "")) + "</div>" +
      "<div>" + esc(v.note || "") + "</div>" +
      "<div class='muted'>" +
        (v.template_name ? "template: " + esc(v.template_name) + "<br>" : "") +
        (v.query_config_id ? "query: " + esc(v.query_config_id) + "<br>" : "") +
        (v.mapping_version_id ? "mapping: " + esc(v.mapping_version_id) : "") +
      "</div>" +
    "</div>";
  }).join("");

  return "<div id='" + attr(outputId) + "'>" + versionRows + "</div>";
}

function renderReportDefinitions(items) {
  const el = document.getElementById("reportDefinitions");

  if (!items.length) {
    el.innerHTML = "<p class='muted'>該当なし</p>";
    return;
  }

  const rows = items.map(item => {
    const isArchived = definitionStatus(item) === "archived";
    const statusClass = isArchived ? "status-disabled" : "status-active";
    const statusText = definitionStatus(item);
    const storageLines = [
      item.gcs_prefix ? "GCS: " + esc(item.gcs_prefix) : "",
      item.drive_folder_name ? "Drive: " + esc(item.drive_folder_name) : ""
    ].filter(Boolean).join("<br>");
    const ownerLines = [
      item.owner ? esc(item.owner) : "",
      item.primary_operator ? "operator: " + esc(item.primary_operator) : ""
    ].filter(Boolean).join("<br>");

    const reportId = item.report_id || "";
    return "<tr>" +
      "<td class='definition-id-cell'><code>" + esc(reportId) + "</code></td>" +
      "<td class='definition-name-cell'><strong>" + esc(item.name || "") + "</strong><br><span class='muted'>" + esc(item.customer_name || "") + "</span></td>" +
      "<td><span class='status-pill " + statusClass + "'>" + esc(statusText) + "</span></td>" +
      "<td>v" + esc(item.current_version || "-") + "<br><span class='muted'>" + esc(item.version_count || 0) + " versions</span>" + renderDefinitionVersions(item) + "</td>" +
      "<td class='definition-storage-cell'>" + (storageLines || "<span class='muted'>-</span>") + "</td>" +
      "<td>" + (ownerLines || "<span class='muted'>-</span>") + "</td>" +
      "<td>" + esc(formatDateTime(item.updated_at || item.created_at || "")) + "<div class='row-actions' style='margin-top:8px;'><button class='small secondary' onclick=\"fillDefinitionForm('" + attr(reportId) + "')\">編集</button></div></td>" +
    "</tr>";
  }).join("");

  el.innerHTML =
    "<p class='muted'>loaded definitions: " + items.length + "件</p>" +
    "<div class='table-wrap'><table class='definition-table'>" +
    "<thead><tr><th>report_id</th><th>name</th><th>status</th><th>current</th><th>保存先</th><th>担当</th><th>更新日時</th></tr></thead>" +
    "<tbody>" + rows + "</tbody></table></div>";
}

function deliveryStatus(item) {
  if (item.active === true) return "active";
  if (item.active === false) return "disabled";
  return "unknown";
}

function deliverySearchText(item) {
  const versions = item.versions || [];
  const fileNames = versions.map(v => v.file_name || "").join(" ");
  return [
    item.delivery_id,
    item.customer_name,
    item.report_month,
    item.download_url,
    item.public_download_url,
    fileNames
  ].join(" ").toLowerCase();
}

function renderDeliveriesFromState() {
  const q = (document.getElementById("deliverySearch").value || "").toLowerCase().trim();
  const status = document.getElementById("deliveryStatusFilter").value;

  const filtered = deliveryItems.filter(item => {
    if (status !== "all" && deliveryStatus(item) !== status) {
      return false;
    }
    if (q && !deliverySearchText(item).includes(q)) {
      return false;
    }
    return true;
  });

  renderDeliveries(filtered);
  updateSummary();
}

function renderVersions(delivery) {
  const versions = delivery.versions || [];
  if (!versions.length) {
    return "<span class='muted'>versionなし</span>";
  }

  return versions.map(v => {
    const current = Number(v.version) === Number(delivery.current_version);
    const gcsUri = v.gcs_uri || "";
    return "<div class='version-panel' style='margin-bottom:8px;'>" +
      "<div><strong>v" + esc(v.version) + "</strong>" + (current ? " <span class='status-pill status-active'>current</span>" : "") + "</div>" +
      "<div class='muted'>" + esc(formatDateTime(v.created_at || "")) + "</div>" +
      "<div><code>" + esc(v.file_name || "") + "</code></div>" +
      "<details class='delivery-uri-toggle'><summary>GCS URIを表示</summary><code>" + esc(gcsUri) + "</code></details>" +
      "</div>";
  }).join("");
}

function versionButtonText(deliveryId) {
  return versionInProgress[deliveryId] ? "更新中..." : "version追加";
}

async function loadDeliveries() {
  const el = document.getElementById("deliveries");
  el.innerHTML = "<p class='muted'>loading deliveries...</p>";

  try {
    const data = await api("/deliveries?limit=100");
    deliveryItems = data.items || [];
    renderDeliveriesFromState();
  } catch (e) {
    el.innerHTML = "<p style='color:#c73535'>" + esc(e.message) + "</p>";
  }
}

function renderDeliveries(items) {
  const el = document.getElementById("deliveries");

  if (!items.length) {
    el.innerHTML = "<p class='muted'>該当なし</p>";
    return;
  }

  const rows = items.map(item => {
    const isActive = item.active === true;
    const statusClass = isActive ? "status-active" : "status-disabled";
    const statusText = isActive ? "active" : "disabled";
    const toggleAction = isActive ? "disable" : "enable";
    const toggleLabel = isActive ? "停止" : "再開";
    const url = item.public_download_url || item.download_url || "";
    const outputId = "versionOutput_" + item.delivery_id;
    const fileInputId = "versionFile_" + item.delivery_id;
    const overwriteId = "overwrite_" + item.delivery_id;

    return "<tr>" +
      "<td class='delivery-id-cell'><code>" + esc(item.delivery_id || "") + "</code></td>" +
      "<td class='delivery-meta-cell'>" + esc(item.customer_name || "") + "<br><span class='muted'>" + esc(item.report_month || "") + "</span></td>" +
      "<td class='delivery-status-cell'><span class='status-pill " + statusClass + "'>" + statusText + "</span></td>" +
      "<td class='delivery-url-cell'><a class='delivery-url-link' href='" + esc(url) + "' target='_blank' rel='noopener noreferrer'>" + esc(url) + "</a></td>" +
      "<td>v" + esc(item.current_version || "") + "</td>" +
      "<td class='delivery-version-cell'>" + renderVersions(item) +
        "<div class='field' style='margin-top:10px;'><label>新規保存ファイル名 .xlsx</label><input id='" + attr(fileInputId) + "' placeholder='例: ダウンロード数入力シート_260522_ICE入力済み_plus.xlsx'></div>" +
        "<div class='field'><label><input type='checkbox' id='" + attr(overwriteId) + "'> overwrite current file name を使う</label></div>" +
        "<div class='row-actions'>" +
          "<button class='small' onclick=\"addVersion('" + attr(item.delivery_id) + "')\" id='versionBtn_" + attr(item.delivery_id) + "'>" + esc(versionButtonText(item.delivery_id)) + "</button>" +
          "<button class='small secondary' onclick=\"loadGcsFiles('" + attr(fileInputId) + "')\">GCS参照</button>" +
        "</div>" +
        "<pre id='" + attr(outputId) + "' style='margin-top:8px;'>待機中</pre>" +
      "</td>" +
      "<td class='delivery-actions-cell'><div class='row-actions'>" +
        "<button class='small secondary' onclick=\"copyText('" + attr(url) + "')\">URLコピー</button>" +
        "<button class='small secondary' onclick=\"loadLogs('" + attr(item.delivery_id) + "')\">ログ</button>" +
        "<button class='small danger' onclick=\"toggleDelivery('" + attr(item.delivery_id) + "', '" + attr(toggleAction) + "')\">" + esc(toggleLabel) + "</button>" +
      "</div></td>" +
    "</tr>";
  }).join("");

  el.innerHTML =
    "<p class='muted'>loaded deliveries: " + items.length + "件</p>" +
    "<div class='table-wrap'><table class='delivery-table'>" +
    "<thead><tr><th>delivery_id</th><th>顧客 / 月</th><th>状態</th><th>配布URL</th><th>current</th><th>versions</th><th>操作</th></tr></thead>" +
    "<tbody>" + rows + "</tbody></table></div>";
}

async function toggleDelivery(deliveryId, action) {
  try {
    await api("/deliveries/" + encodeURIComponent(deliveryId) + "/" + action, {
      method: "POST"
    });
    showToast(action === "disable" ? "停止しました" : "再開しました");
    await loadDeliveries();
  } catch (e) {
    alert(e.message);
  }
}

async function addVersion(deliveryId) {
  if (versionInProgress[deliveryId]) {
    return;
  }

  versionInProgress[deliveryId] = true;

  const button = document.getElementById("versionBtn_" + deliveryId);
  const fileInput = document.getElementById("versionFile_" + deliveryId);
  const overwrite = document.getElementById("overwrite_" + deliveryId);
  const output = document.getElementById("versionOutput_" + deliveryId);

  if (button) {
    button.disabled = true;
    button.textContent = versionButtonText(deliveryId);
  }

  if (output) {
    output.textContent = "version追加中...";
  }

  try {
    const payload = {
      output_filename: fileInput ? fileInput.value : "",
      overwrite: overwrite ? overwrite.checked : false
    };

    const data = await api("/deliveries/" + encodeURIComponent(deliveryId) + "/versions", {
      method: "POST",
      body: JSON.stringify(payload)
    });

    if (output) {
      output.textContent = "version追加完了\n" + JSON.stringify(data.result || data, null, 2);
    }
    showToast("versionを追加しました");
    await loadDeliveries();
    await loadGcsFiles();
  } catch (e) {
    if (output) {
      output.textContent = "version追加に失敗しました\n" + e.message;
    }
  } finally {
    versionInProgress[deliveryId] = false;
    if (button) {
      button.disabled = false;
      button.textContent = versionButtonText(deliveryId);
    }
  }
}

function updateSummary() {
  const total = deliveryItems.length;
  const active = deliveryItems.filter(item => item.active === true).length;
  const disabled = deliveryItems.filter(item => item.active === false).length;
  const shownLogs = logItems.length;
  const definitions = reportDefinitionItems.length;

  document.getElementById("summaryTotal").textContent = String(total);
  document.getElementById("summaryActive").textContent = String(active);
  document.getElementById("summaryDisabled").textContent = String(disabled);
  document.getElementById("summaryDefinitionsLogs").textContent = String(definitions) + " / " + String(shownLogs);
}

async function loadLogs(deliveryId = "") {
  const logsEl = document.getElementById("logs");
  logsEl.innerHTML = "<p class='muted'>loading logs...</p>";

  try {
    const path = deliveryId
      ? "/download-logs?delivery_id=" + encodeURIComponent(deliveryId)
      : "/download-logs";

    const data = await api(path);
    logItems = data.items || data.logs || [];
    renderLogsFromState(deliveryId);
    updateSummary();

  } catch (e) {
    logsEl.innerHTML =
      "<p style='color:#c73535'>" + esc(e.message) + "</p>";
  }
}

function logSearchText(item) {
  return [
    item.downloaded_at,
    item.delivery_id,
    item.customer_name,
    item.report_month,
    item.email,
    item.version,
    item.file_name
  ].join(" ").toLowerCase();
}

function renderLogsFromState(deliveryId = "") {
  const q = (document.getElementById("logSearch").value || "").toLowerCase().trim();
  const filtered = logItems.filter(item => !q || logSearchText(item).includes(q));
  renderLogs(filtered, deliveryId);
  updateSummary();
}

function renderLogs(items, deliveryId = "") {
  const el = document.getElementById("logs");

  if (!items.length) {
    el.innerHTML = "<p class='muted'>該当なし</p>";
    return;
  }

  const rows = items.map(item =>
    "<tr>" +
      "<td>" + esc(formatDateTime(item.downloaded_at || "")) + "</td>" +
      "<td><code>" + esc(item.delivery_id || "") + "</code></td>" +
      "<td>" + esc(item.customer_name || "") + "<br><span class='muted'>" + esc(item.report_month || "") + "</span></td>" +
      "<td>" + esc(item.email || "") + "</td>" +
      "<td><span class='status-pill status-warning'>v" + esc(item.version || "") + "</span></td>" +
      "<td><code>" + esc(item.file_name || "") + "</code></td>" +
    "</tr>"
  ).join("");

  el.innerHTML =
    "<p class='muted'>loaded logs: " + items.length + "件 / delivery_id=" + esc(deliveryId || "all") + "</p>" +
    "<div class='table-wrap'><table>" +
    "<thead><tr><th>日時</th><th>delivery_id</th><th>顧客/月</th><th>email</th><th>version</th><th>file</th></tr></thead>" +
    "<tbody>" + rows + "</tbody></table></div>";
}

async function loadAll() {
  await loadReportDefinitions();
  await loadDeliveries();
  await loadLogs();
  await loadGcsFiles();
}

loadReportDefinitions();
loadDeliveries();
loadLogs();

</script>
</body>
</html>
"""


@app.get("/admin")
def admin_ui():
    return render_admin_ui()


@app.get("/gcs-files")
def list_gcs_files():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    bucket_name = os.environ.get("BUCKET_NAME")
    if not bucket_name:
        return jsonify({"error": "BUCKET_NAME is required"}), 400

    prefix = request.args.get("prefix") or "reports/plus/"
    limit = int(request.args.get("limit", "50"))

    client = storage.Client()
    blobs = client.list_blobs(bucket_name, prefix=prefix)

    files = []
    for blob in blobs:
        if not blob.name.endswith(".xlsx"):
            continue

        files.append({
            "name": blob.name,
            "gcs_uri": f"gs://{bucket_name}/{blob.name}",
            "updated": blob.updated.isoformat() if blob.updated else "",
            "size": blob.size,
        })

    files.sort(key=lambda x: x.get("updated") or "", reverse=True)

    return jsonify({"items": files[:limit]})


@app.post("/generate")
def generate():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}

    project_id = payload.get("project_id") or _bigquery_project_id()
    bucket_name = payload.get("bucket_name") or os.environ.get("BUCKET_NAME")
    object_prefix = payload.get("object_prefix") or os.environ.get("OBJECT_PREFIX", "reports/plus")
    today_text = payload.get("today")
    output_filename = payload.get("output_filename") or payload.get("file_name")

    if not project_id:
        _log_admin_audit_event(
            action="generate_report",
            result="failure",
            target_type="report",
            status_code=400,
            reason="project_required",
        )
        return jsonify({"error": "BIGQUERY_PROJECT_ID or PROJECT_ID is required"}), 400

    if not bucket_name:
        _log_admin_audit_event(
            action="generate_report",
            result="failure",
            target_type="report",
            status_code=400,
            reason="bucket_required",
        )
        return jsonify({"error": "BUCKET_NAME is required"}), 400

    today = datetime.strptime(today_text, "%Y-%m-%d").date() if today_text else None
    template_context = None

    try:
        template_context = _resolve_generation_template(payload)
    except RuntimeTemplateError as exc:
        _log_admin_audit_event(
            action="generate_report",
            result="failure",
            target_type="report",
            status_code=exc.status_code,
            reason=exc.reason,
            detail={"report_id": str(payload.get("report_id") or "").strip()},
        )
        return jsonify({"error": exc.message}), exc.status_code

    try:
        result = generate_report(
            project_id=project_id,
            bucket_name=bucket_name,
            object_prefix=object_prefix,
            template_path=template_context["template_path"],
            today=today,
            output_filename=output_filename,
        )
    finally:
        _cleanup_runtime_template(template_context.get("local_path") if template_context else None)

    if template_context and template_context.get("runtime_template"):
        result["report_definition_template"] = template_context["runtime_template"]

    _log_admin_audit_event(
        action="generate_report",
        result="success",
        target_type="report",
        status_code=200,
        detail={
            "bucket_name": bucket_name,
            "object_prefix": object_prefix,
            "output_filename": output_filename,
            "gcs_uri": result.get("gcs_uri"),
            "item_count": len(result.get("items", [])) if isinstance(result.get("items"), list) else None,
            "report_definition_template": (
                template_context.get("runtime_template") if template_context else None
            ),
        },
    )

    return jsonify({"items": result["items"] if "items" in result else [result], "result": result})


@app.post("/deliveries")
def create_delivery():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}

    customer_name = payload.get("customer_name")
    report_month = payload.get("report_month")
    gcs_uri = payload.get("gcs_uri")
    output_filename = payload.get("output_filename") or payload.get("file_name")
    template_context = None

    if not customer_name:
        _log_admin_audit_event(
            action="delivery_create",
            result="failure",
            target_type="delivery",
            status_code=400,
            reason="customer_name_required",
        )
        return jsonify({"error": "customer_name is required"}), 400

    if not report_month:
        _log_admin_audit_event(
            action="delivery_create",
            result="failure",
            target_type="delivery",
            status_code=400,
            reason="report_month_required",
            detail={"customer_name": customer_name},
        )
        return jsonify({"error": "report_month is required"}), 400

    if not gcs_uri:
        project_id = payload.get("project_id") or _bigquery_project_id()
        bucket_name = payload.get("bucket_name") or os.environ.get("BUCKET_NAME")
        object_prefix = payload.get("object_prefix") or os.environ.get("OBJECT_PREFIX", "reports/plus")
        today_text = payload.get("today")

        if not project_id:
            _log_admin_audit_event(
                action="delivery_create",
                result="failure",
                target_type="delivery",
                status_code=400,
                reason="project_required",
                detail={"customer_name": customer_name, "report_month": report_month},
            )
            return jsonify({"error": "BIGQUERY_PROJECT_ID or PROJECT_ID is required"}), 400

        if not bucket_name:
            _log_admin_audit_event(
                action="delivery_create",
                result="failure",
                target_type="delivery",
                status_code=400,
                reason="bucket_required",
                detail={"customer_name": customer_name, "report_month": report_month},
            )
            return jsonify({"error": "BUCKET_NAME is required"}), 400

        today = datetime.strptime(today_text, "%Y-%m-%d").date() if today_text else None
        try:
            template_context = _resolve_generation_template(payload)
        except RuntimeTemplateError as exc:
            _log_admin_audit_event(
                action="delivery_create",
                result="failure",
                target_type="delivery",
                status_code=exc.status_code,
                reason=exc.reason,
                detail={
                    "customer_name": customer_name,
                    "report_month": report_month,
                    "report_id": str(payload.get("report_id") or "").strip(),
                },
            )
            return jsonify({"error": exc.message}), exc.status_code

        try:
            generated = generate_report(
                project_id=project_id,
                bucket_name=bucket_name,
                object_prefix=object_prefix,
                template_path=template_context["template_path"],
                today=today,
                output_filename=output_filename,
            )
        finally:
            _cleanup_runtime_template(template_context.get("local_path") if template_context else None)

        if template_context and template_context.get("runtime_template"):
            generated["report_definition_template"] = template_context["runtime_template"]

        gcs_uri = generated.get("gcs_uri")

        if not gcs_uri:
            _log_admin_audit_event(
                action="delivery_create",
                result="failure",
                target_type="delivery",
                status_code=500,
                reason="gcs_uri_not_generated",
                detail={"customer_name": customer_name, "report_month": report_month},
            )
            return jsonify({
                "error": "gcs_uri was not generated",
                "generated": generated,
            }), 500

    result = create_delivery_record(
        customer_name=customer_name,
        report_month=report_month,
        gcs_uri=gcs_uri,
        allowed_domains=payload.get("allowed_domains") or [],
        allowed_emails=payload.get("allowed_emails") or [],
        expires_days=int(
            payload.get(
                "expires_days",
                os.environ.get("DEFAULT_EXPIRES_DAYS", "7")
            )
        ),
        version_note=payload.get("version_note"),
    )

    _log_admin_audit_event(
        action="delivery_create",
        result="success",
        target_type="delivery",
        target_id=result.get("delivery_id", ""),
        status_code=201,
        detail={
            "customer_name": customer_name,
            "report_month": report_month,
            "gcs_uri": gcs_uri,
            "expires_days": int(
                payload.get(
                    "expires_days",
                    os.environ.get("DEFAULT_EXPIRES_DAYS", "7")
                )
            ),
            "allowed_domain_count": len(payload.get("allowed_domains") or []),
            "allowed_email_count": len(payload.get("allowed_emails") or []),
            "report_definition_template": (
                template_context.get("runtime_template") if template_context else None
            ),
        },
    )

    response_payload = {
        "items": [result],
        "result": result,
        "download_url": result.get("download_url"),
        "public_download_url": result.get("public_download_url"),
        "token": result.get("token"),
        "delivery_id": result.get("delivery_id"),
    }
    if template_context and template_context.get("runtime_template"):
        response_payload["report_definition_template"] = template_context["runtime_template"]

    return jsonify(response_payload), 201


@app.get("/deliveries")
def list_deliveries():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    report_month = request.args.get("report_month")
    active_text = request.args.get("active")
    active = None

    if active_text is not None:
        active = active_text.lower() in ("1", "true", "yes", "y")

    limit = int(request.args.get("limit", "100"))

    result = list_delivery_records(
        report_month=report_month,
        active=active,
        limit=limit,
    )

    return jsonify({"items": result})


@app.get("/report-definitions")
def report_definitions():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    limit = int(request.args.get("limit", "100"))
    result = list_report_definitions(limit=limit)

    return jsonify({"items": result})


def _log_report_definition_action(action: str, result: str, report_id: str, status_code: int) -> None:
    logging.warning(
        "ICE_REPORT_REPORT_DEFINITION_ACTION action=%s result=%s report_id=%s status_code=%s",
        action,
        result,
        report_id,
        status_code,
    )


def _safe_uploaded_filename(filename: str) -> str:
    return (filename or "").replace("\\", "/").rsplit("/", 1)[-1]


def _preview_xlsx_template_bytes(
    data: bytes,
    filename: str,
    *,
    max_bytes: int = TEMPLATE_PREVIEW_MAX_BYTES,
) -> dict:
    safe_filename = _safe_uploaded_filename(filename)

    if not safe_filename.lower().endswith(".xlsx"):
        raise ValueError("template file must end with .xlsx")
    if not data:
        raise ValueError("template file is empty")
    if len(data) > max_bytes:
        raise ValueError("template file is too large")

    try:
        from openpyxl import load_workbook as openpyxl_load_workbook

        workbook = openpyxl_load_workbook(BytesIO(data), read_only=False, data_only=False)
    except Exception as exc:
        raise ValueError("invalid xlsx template") from exc

    try:
        sheets = []
        for worksheet in workbook.worksheets:
            tables = getattr(worksheet, "tables", {}) or {}
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "table_count": len(tables),
                    "state": worksheet.sheet_state,
                }
            )

        return {
            "file_name": safe_filename,
            "size_bytes": len(data),
            "sha256": hashlib.sha256(data).hexdigest(),
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()


@app.post("/report-definitions")
def create_report_definition_route():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}
    report_id = str(payload.get("report_id") or "")

    try:
        result = create_report_definition(payload)
    except ValueError as exc:
        _log_report_definition_action("report_definition_create", "failure", "", 400)
        return jsonify({"error": str(exc)}), 400

    _log_report_definition_action("report_definition_create", "success", result.get("report_id", ""), 201)
    return jsonify({"item": result, "result": result}), 201


@app.get("/report-definitions/<report_id>")
def report_definition_detail(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    try:
        result = get_report_definition(report_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404

    return jsonify({
        "item": result,
        "result": result,
    })


@app.post("/report-definitions/<report_id>/query-mapping-preview")
def preview_report_definition_query_mapping(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}
    project_id = payload.get("project_id") or _bigquery_project_id()
    if not project_id:
        _log_report_definition_action("query_mapping_preview", "failure", "", 400)
        return jsonify({"error": "BIGQUERY_PROJECT_ID or PROJECT_ID is required"}), 400

    try:
        get_report_definition(report_id)
    except ValueError as exc:
        _log_report_definition_action("query_mapping_preview", "failure", "", 404)
        return jsonify({"error": str(exc)}), 404

    try:
        preview = preview_default_query_mapping(project_id)
    except Exception:
        logging.error("ICE_REPORT_QUERY_MAPPING_PREVIEW_FAILED report_id=%s", report_id)
        _log_report_definition_action("query_mapping_preview", "failure", "", 500)
        return jsonify({"error": "query mapping preview failed"}), 500

    preview["report_id"] = report_id
    _log_report_definition_action("query_mapping_preview", "success", report_id, 200)
    return jsonify({"preview": preview, "result": preview})


@app.post("/report-definitions/<report_id>/template-preview")
def preview_report_definition_template(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    try:
        get_report_definition(report_id)
    except ValueError as exc:
        _log_report_definition_action("template_preview", "failure", "", 404)
        return jsonify({"error": str(exc)}), 404

    upload = request.files.get("template_file")
    if upload is None:
        _log_report_definition_action("template_preview", "failure", "", 400)
        return jsonify({"error": "template_file is required"}), 400

    data = upload.read(TEMPLATE_PREVIEW_MAX_BYTES + 1)

    try:
        preview = _preview_xlsx_template_bytes(data, upload.filename or "")
    except ValueError as exc:
        status_code = 413 if "too large" in str(exc) else 400
        _log_report_definition_action("template_preview", "failure", "", status_code)
        return jsonify({"error": str(exc)}), status_code

    preview["report_id"] = report_id
    _log_report_definition_action("template_preview", "success", report_id, 200)
    return jsonify({"preview": preview, "result": preview})


@app.post("/report-definitions/<report_id>/template-publish")
def publish_report_definition_template_route(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    bucket_name = os.environ.get("TEMPLATE_BUCKET_NAME") or os.environ.get("BUCKET_NAME")
    object_prefix = os.environ.get("TEMPLATE_OBJECT_PREFIX", "report-templates")
    if not bucket_name:
        _log_report_definition_action("template_publish", "failure", "", 400)
        return jsonify({"error": "TEMPLATE_BUCKET_NAME or BUCKET_NAME is required"}), 400

    upload = request.files.get("template_file")
    if upload is None:
        _log_report_definition_action("template_publish", "failure", "", 400)
        return jsonify({"error": "template_file is required"}), 400

    data = upload.read(TEMPLATE_PREVIEW_MAX_BYTES + 1)

    try:
        preview = _preview_xlsx_template_bytes(data, upload.filename or "")
        result = publish_report_definition_template(
            report_id,
            template_bytes=data,
            preview=preview,
            bucket_name=bucket_name,
            object_prefix=object_prefix,
            note=request.form.get("note") or "",
        )
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 413 if "too large" in str(exc) else 400
        _log_report_definition_action("template_publish", "failure", "", status_code)
        return jsonify({"error": str(exc)}), status_code
    except Exception:
        logging.error("ICE_REPORT_TEMPLATE_PUBLISH_FAILED")
        _log_report_definition_action("template_publish", "failure", "", 500)
        return jsonify({"error": "template publish failed"}), 500

    _log_report_definition_action("template_publish", "success", report_id, 201)
    return jsonify({**result, "result": result.get("template")}), 201


@app.post("/report-definitions/<report_id>/template-rollback")
def rollback_report_definition_template_route(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}
    try:
        version = int(payload.get("version") or 0)
        result = rollback_report_definition_template(report_id, version)
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 400
        _log_report_definition_action("template_rollback", "failure", "", status_code)
        return jsonify({"error": str(exc)}), status_code
    except Exception:
        logging.error("ICE_REPORT_TEMPLATE_ROLLBACK_FAILED")
        _log_report_definition_action("template_rollback", "failure", "", 500)
        return jsonify({"error": "template rollback failed"}), 500

    _log_report_definition_action("template_rollback", "success", report_id, 200)
    return jsonify({"item": result, "result": result})


@app.patch("/report-definitions/<report_id>")
def update_report_definition_route(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}

    try:
        result = update_report_definition(report_id, payload)
    except ValueError as exc:
        status_code = 404 if "not found" in str(exc) else 400
        _log_report_definition_action("report_definition_update", "failure", "", status_code)
        return jsonify({"error": str(exc)}), status_code

    _log_report_definition_action("report_definition_update", "success", report_id, 200)
    return jsonify({"item": result, "result": result})


@app.post("/report-definitions/<report_id>/archive")
def archive_report_definition_route(report_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    try:
        result = archive_report_definition(report_id)
    except ValueError as exc:
        _log_report_definition_action("report_definition_archive", "failure", "", 404)
        return jsonify({"error": str(exc)}), 404

    _log_report_definition_action("report_definition_archive", "success", report_id, 200)
    return jsonify({"item": result, "result": result})


@app.post("/deliveries/<delivery_id>/versions")
def add_version(delivery_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    payload = request.get_json(silent=True) or {}

    project_id = payload.get("project_id") or _bigquery_project_id()
    bucket_name = payload.get("bucket_name") or os.environ.get("BUCKET_NAME")
    object_prefix = payload.get("object_prefix") or os.environ.get("OBJECT_PREFIX", "reports/plus")
    today_text = payload.get("today")
    output_filename = payload.get("output_filename") or payload.get("file_name")
    overwrite = bool(payload.get("overwrite", False))

    if not project_id:
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=400,
            reason="project_required",
        )
        return jsonify({"error": "BIGQUERY_PROJECT_ID or PROJECT_ID is required"}), 400

    if not bucket_name:
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=400,
            reason="bucket_required",
        )
        return jsonify({"error": "BUCKET_NAME is required"}), 400

    if overwrite:
        deliveries = list_delivery_records(limit=500)
        target = next(
            (item for item in deliveries if item.get("delivery_id") == delivery_id),
            None,
        )

        if not target:
            _log_admin_audit_event(
                action="delivery_version_add",
                result="failure",
                target_type="delivery",
                target_id=delivery_id,
                status_code=404,
                reason="delivery_not_found",
                detail={"overwrite": overwrite},
            )
            return jsonify({"error": "delivery_id not found"}), 404

        versions = target.get("versions") or []
        current_version = target.get("current_version")
        current = next(
            (v for v in versions if v.get("version") == current_version),
            None,
        )

        if not current:
            _log_admin_audit_event(
                action="delivery_version_add",
                result="failure",
                target_type="delivery",
                target_id=delivery_id,
                status_code=404,
                reason="current_version_not_found",
                detail={"overwrite": overwrite, "current_version": current_version},
            )
            return jsonify({"error": "current version not found"}), 404

        output_filename = current.get("file_name")

    if not output_filename:
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=400,
            reason="output_filename_required",
            detail={"overwrite": overwrite},
        )
        return jsonify({"error": "output_filename is required"}), 400

    if not output_filename.endswith(".xlsx"):
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=400,
            reason="invalid_output_filename_extension",
            detail={"overwrite": overwrite, "output_filename": output_filename},
        )
        return jsonify({"error": "output_filename must end with .xlsx"}), 400

    today = datetime.strptime(today_text, "%Y-%m-%d").date() if today_text else None

    generated = generate_report(
        project_id=project_id,
        bucket_name=bucket_name,
        object_prefix=object_prefix,
        template_path=Path(os.environ.get("TEMPLATE_PATH", str(DEFAULT_TEMPLATE))),
        today=today,
        output_filename=output_filename,
    )

    gcs_uri = generated.get("gcs_uri")
    if not gcs_uri:
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=500,
            reason="gcs_uri_not_generated",
            detail={"overwrite": overwrite, "output_filename": output_filename},
        )
        return jsonify({
            "error": "gcs_uri was not generated",
            "generated": generated,
        }), 500

    try:
        result = add_delivery_version(
            delivery_id=delivery_id,
            gcs_uri=gcs_uri,
            note=payload.get("note"),
            make_current=bool(payload.get("make_current", True)),
        )
    except ValueError as exc:
        _log_admin_audit_event(
            action="delivery_version_add",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=404,
            reason=str(exc),
            detail={"overwrite": overwrite, "output_filename": output_filename},
        )
        return jsonify({"error": str(exc)}), 404

    _log_admin_audit_event(
        action="delivery_version_add",
        result="success",
        target_type="delivery",
        target_id=delivery_id,
        status_code=201,
        detail={
            "version": result.get("version"),
            "current_version": result.get("current_version"),
            "gcs_uri": gcs_uri,
            "output_filename": output_filename,
            "overwrite": overwrite,
            "make_current": bool(payload.get("make_current", True)),
        },
    )

    return jsonify({
        "items": [result],
        "result": result,
        "generated": generated,
    }), 201


@app.post("/deliveries/<delivery_id>/disable")
def disable_delivery(delivery_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    try:
        result = set_delivery_active(delivery_id, False)
    except ValueError as exc:
        _log_admin_audit_event(
            action="delivery_disable",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=404,
            reason=str(exc),
        )
        return jsonify({"error": str(exc)}), 404

    _log_admin_audit_event(
        action="delivery_disable",
        result="success",
        target_type="delivery",
        target_id=delivery_id,
        status_code=200,
        detail={"active": result.get("active")},
    )

    return jsonify({
        "items": [result],
        "result": result,
    })


@app.post("/deliveries/<delivery_id>/enable")
def enable_delivery(delivery_id: str):
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    try:
        result = set_delivery_active(delivery_id, True)
    except ValueError as exc:
        _log_admin_audit_event(
            action="delivery_enable",
            result="failure",
            target_type="delivery",
            target_id=delivery_id,
            status_code=404,
            reason=str(exc),
        )
        return jsonify({"error": str(exc)}), 404

    _log_admin_audit_event(
        action="delivery_enable",
        result="success",
        target_type="delivery",
        target_id=delivery_id,
        status_code=200,
        detail={"active": result.get("active")},
    )

    return jsonify({
        "items": [result],
        "result": result,
    })


@app.get("/download-logs")
def list_download_logs():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    delivery_id = request.args.get("delivery_id")
    email = request.args.get("email")
    limit = int(request.args.get("limit", "100"))

    result = list_download_log_records(
        delivery_id=delivery_id,
        email=email,
        limit=limit,
    )

    return jsonify({
        "items": result,
        "logs": result,
    })


@app.post("/internal/cleanup")
def cleanup_expired_deliveries():
    ok, error_response = _check_admin()
    if not ok:
        return error_response

    collection_name = os.environ.get("DELIVERIES_COLLECTION", "deliveries")
    now = datetime.now(timezone.utc)
    updated = []
    count = 0

    try:
        db = firestore.Client()
        query = (
            db.collection(collection_name)
            .where("active", "==", True)
            .where("expires_at", "<", now)
        )
        batch = db.batch()

        for doc in query.stream():
            ref = db.collection(collection_name).document(doc.id)
            batch.update(
                ref,
                {
                    "active": False,
                    "updated_at": now,
                    "cleanup_reason": "expired",
                    "cleanup_at": now,
                },
            )
            updated.append(doc.id)
            count += 1

            if count % 400 == 0:
                batch.commit()
                batch = db.batch()

        if count % 400 != 0:
            batch.commit()
    except Exception as exc:
        _log_admin_audit_event(
            action="cleanup_expired_deliveries",
            result="failure",
            target_type="delivery",
            status_code=500,
            reason=exc.__class__.__name__,
            detail={"collection_name": collection_name, "updated_count": count},
        )
        raise

    _log_admin_audit_event(
        action="cleanup_expired_deliveries",
        result="success",
        target_type="delivery",
        status_code=200,
        detail={
            "collection_name": collection_name,
            "updated_count": count,
            "updated_delivery_ids": updated,
        },
    )

    return jsonify({
        "status": "ok",
        "updated_count": count,
        "updated_delivery_ids": updated,
        "cleanup_at": now.isoformat(),
    })


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _otp_collection_name() -> str:
    return os.environ.get("OTP_COLLECTION", "otp_challenges")


def _download_sessions_collection_name() -> str:
    return os.environ.get("DOWNLOAD_SESSIONS_COLLECTION", "download_sessions")


def _otp_hash_secret() -> str:
    secret = (
        os.environ.get("OTP_HASH_SECRET")
        or os.environ.get("SECRET_KEY")
        or os.environ.get("ADMIN_API_KEY")
    )
    if not secret:
        logging.warning(
            "OTP_HASH_SECRET, SECRET_KEY, and ADMIN_API_KEY are not set. "
            "Using PROJECT_ID fallback for OTP hashing. "
            "Set OTP_HASH_SECRET before production use."
        )
        secret = os.environ.get("PROJECT_ID", "ice-report-local-dev")
    return secret


def _hash_value(value: str) -> str:
    return hmac.new(
        _otp_hash_secret().encode("utf-8"),
        value.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def _log_fingerprint(value: str) -> str:
    if not value:
        return ""

    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:16]


def _admin_actor_context() -> dict:
    if _admin_iap_auth_enabled():
        email = _request_iap_email()
        if email:
            return {
                "actor_type": "iap_user" if email in _admin_iap_allowed_emails() else "iap_user_denied",
                "admin_key_fingerprint": "",
                "iap_email_hash": _log_fingerprint(email),
            }

    return {
        "actor_type": "admin_key",
        "admin_key_fingerprint": _log_fingerprint(request.headers.get("X-Admin-Key", "")),
        "iap_email_hash": "",
    }


def _normalize_email(email: str) -> str:
    return (email or "").strip().lower()


def _cookie_name(token: str) -> str:
    safe = "".join(c if c.isalnum() else "_" for c in token)
    return f"ice_dl_session_{safe}"


def _int_env(name: str, default_value: int) -> int:
    try:
        return int(os.environ.get(name, str(default_value)))
    except ValueError:
        return default_value


def _bool_env(name: str, default_value: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default_value
    return value.lower() in ("1", "true", "yes", "y", "on")


def _format_page_datetime(value) -> str:
    if not value:
        return ""

    if isinstance(value, datetime):
        return value.isoformat()

    return str(value)


def _render_selected_report_summary(delivery: dict | None) -> str:
    if not delivery:
        return ""

    current_version = delivery.get("current_version")
    current = {}
    for version in delivery.get("versions") or []:
        if version.get("version") == current_version:
            current = version
            break

    active_text = "active" if delivery.get("active") else "disabled"
    rows = [
        ("顧客", delivery.get("customer_name") or "-"),
        ("対象月", delivery.get("report_month") or "-"),
        ("current", f"v{current_version}" if current_version else "-"),
        ("file", current.get("file_name") or "-"),
        ("期限", _format_page_datetime(delivery.get("expires_at")) or "-"),
        ("状態", active_text),
    ]

    row_html = "".join(
        f"<dt>{escape(label)}</dt><dd>{escape(str(value))}</dd>"
        for label, value in rows
    )

    return f"""
    <section class="report-summary" aria-label="選択中レポート">
      <h2>選択中レポート</h2>
      <dl>{row_html}</dl>
    </section>
"""


def _render_otp_page(
    token: str,
    *,
    delivery: dict | None = None,
    email: str = "",
    step: str = "email",
    message: str = "",
    error: str = "",
) -> str:
    token_value = escape(token or "", quote=True)
    email_value = escape(email or "", quote=True)
    message_html = (
        f"<div class='message'>{escape(message)}</div>"
        if message else ""
    )
    error_html = (
        f"<div class='error'>{escape(error)}</div>"
        if error else ""
    )
    report_summary_html = _render_selected_report_summary(delivery)

    if step == "pin":
        form_html = f"""
<form method="post" action="/d/{token_value}/verify-pin">
  <label>メールアドレス</label>
  <input type="email" name="email" value="{email_value}" required readonly>
  <label>PIN</label>
  <input type="text" name="pin" inputmode="numeric" pattern="[0-9]{{6}}" maxlength="6" placeholder="6桁のPIN" required>
  <button type="submit">PINを確認してダウンロードへ進む</button>
</form>
<form method="post" action="/d/{token_value}/request-pin" class="secondary-form">
  <input type="hidden" name="email" value="{email_value}">
  <button type="submit" class="secondary">PINを再発行</button>
</form>
"""
    else:
        form_html = f"""
<form method="post" action="/d/{token_value}/request-pin">
  <label>メールアドレス</label>
  <input type="email" name="email" value="{email_value}" placeholder="you@example.com" required>
  <button type="submit">PINを送信</button>
</form>
"""

    return f"""
<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>ICEレポート ダウンロード</title>
  <style>
    :root {{
      color-scheme: light dark;
      --bg: #f5f7fb;
      --panel: rgba(255, 255, 255, 0.94);
      --text: #172033;
      --muted: #667085;
      --line: #d8dee9;
      --primary: #2457d6;
      --primary-dark: #1c45ab;
      --success: #157347;
      --success-bg: #eaf7ef;
      --danger: #c73535;
      --danger-bg: #fff1f1;
      --radius: 18px;
      --shadow: 0 24px 48px rgba(20, 32, 55, 0.14);
    }}

    @media (prefers-color-scheme: dark) {{
      :root {{
        --bg: #0b1020;
        --panel: rgba(18, 25, 43, 0.92);
        --text: #e6edf7;
        --muted: #9aa8bd;
        --line: #2d3a53;
        --primary: #7aa2ff;
        --primary-dark: #5f8df0;
        --success: #65d99a;
        --success-bg: rgba(101, 217, 154, 0.13);
        --danger: #ff7b7b;
        --danger-bg: rgba(255, 123, 123, 0.13);
        --shadow: 0 24px 52px rgba(0, 0, 0, 0.38);
      }}
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      min-height: 100vh;
      display: grid;
      place-items: center;
      background:
        radial-gradient(circle at top left, rgba(36, 87, 214, 0.10), transparent 34rem),
        radial-gradient(circle at top right, rgba(21, 115, 71, 0.08), transparent 28rem),
        var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.55;
      padding: 24px;
    }}

    .card {{
      width: min(100%, 460px);
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 24px;
    }}

    h1 {{
      margin: 0 0 8px;
      font-size: 22px;
      letter-spacing: -0.02em;
    }}

    p {{
      margin: 0 0 18px;
      color: var(--muted);
      font-size: 14px;
    }}

    label {{
      display: block;
      margin: 14px 0 6px;
      color: var(--muted);
      font-size: 12px;
    }}

    input {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 11px 12px;
      font: inherit;
      background: var(--panel);
      color: var(--text);
    }}

    button {{
      width: 100%;
      margin-top: 18px;
      border: 0;
      border-radius: 10px;
      padding: 11px 12px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      color: #fff;
      background: var(--primary);
    }}

    button:hover {{ background: var(--primary-dark); }}

    button.secondary {{
      color: var(--text);
      background: transparent;
      border: 1px solid var(--line);
    }}

    .secondary-form {{
      margin-top: 8px;
    }}

    .message {{
      border: 1px solid rgba(21, 115, 71, 0.35);
      border-radius: 10px;
      padding: 10px 12px;
      background: var(--success-bg);
      color: var(--success);
      font-size: 13px;
      margin-bottom: 14px;
    }}

    .error {{
      border: 1px solid rgba(199, 53, 53, 0.35);
      border-radius: 10px;
      padding: 10px 12px;
      background: var(--danger-bg);
      color: var(--danger);
      font-size: 13px;
      margin-bottom: 14px;
    }}

    .note {{
      margin-top: 16px;
      color: var(--muted);
      font-size: 12px;
    }}

    .report-summary {{
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 14px;
      margin: 18px 0;
      background: color-mix(in srgb, var(--panel) 88%, var(--primary) 12%);
    }}

    .report-summary h2 {{
      margin: 0 0 10px;
      font-size: 15px;
    }}

    .report-summary dl {{
      display: grid;
      grid-template-columns: 96px 1fr;
      gap: 6px 10px;
      margin: 0;
      font-size: 13px;
    }}

    .report-summary dt {{
      color: var(--muted);
    }}

    .report-summary dd {{
      margin: 0;
      overflow-wrap: anywhere;
    }}
  </style>
</head>
<body>
  <main class="card">
    <h1>ICEレポート ダウンロード</h1>
    <p>許可されたメールアドレス宛に発行されたPINで認証します。</p>
    {report_summary_html}
    {message_html}
    {error_html}
    {form_html}
    <div class="note">PINの有効期限は約10分です。</div>
  </main>
</body>
</html>
"""


def _get_client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()

    return request.remote_addr or ""


def _revoke_existing_challenges(token: str, email: str) -> int:
    now = _now_utc()
    db = firestore.Client()
    query = (
        db.collection(_otp_collection_name())
        .where("token", "==", token)
        .where("email", "==", email)
        .where("used", "==", False)
        .limit(100)
    )

    batch = db.batch()
    count = 0

    for doc in query.stream():
        batch.update(
            db.collection(_otp_collection_name()).document(doc.id),
            {
                "used": True,
                "revoked": True,
                "revoked_at": now,
                "revoked_reason": "superseded_by_new_pin",
            },
        )
        count += 1

    if count:
        batch.commit()

    return count


def _security_events_collection_name() -> str:
    return os.environ.get("SECURITY_EVENTS_COLLECTION", "security_events")


def _admin_audit_logs_collection_name() -> str:
    return os.environ.get("ADMIN_AUDIT_LOGS_COLLECTION", "admin_audit_logs")


def _log_admin_audit_event(
    *,
    action: str,
    result: str,
    target_type: str = "",
    target_id: str = "",
    status_code: int | None = None,
    reason: str = "",
    detail: dict | None = None,
) -> None:
    now = _now_utc()
    actor = _admin_actor_context()
    record = {
        "action": action,
        "result": result,
        "target_type": target_type,
        "target_id": target_id,
        "status_code": status_code,
        "reason": reason,
        "detail": detail or {},
        "actor_type": actor["actor_type"],
        "admin_key_fingerprint": actor["admin_key_fingerprint"],
        "iap_email_hash": actor["iap_email_hash"],
        "path": request.path,
        "method": request.method,
        "ip": _get_client_ip(),
        "user_agent": request.headers.get("User-Agent", ""),
        "created_at": now,
    }

    try:
        db = firestore.Client()
        db.collection(_admin_audit_logs_collection_name()).document().set(record)
    except Exception:
        logging.exception("failed to write admin audit log")

    logging.warning(
        "ICE_REPORT_ADMIN_AUDIT action=%s result=%s target_type=%s target_id=%s status_code=%s reason=%s",
        action,
        result,
        target_type,
        target_id,
        status_code,
        reason,
    )


def _log_security_event(
    *,
    event_type: str,
    token: str = "",
    delivery_id: str = "",
    email: str = "",
    reason: str = "",
    detail: dict | None = None,
) -> None:
    now = _now_utc()
    record = {
        "event_type": event_type,
        "token": token,
        "delivery_id": delivery_id,
        "email": email,
        "reason": reason,
        "detail": detail or {},
        "ip": _get_client_ip(),
        "user_agent": request.headers.get("User-Agent", ""),
        "created_at": now,
    }

    try:
        db = firestore.Client()
        db.collection(_security_events_collection_name()).document().set(record)
    except Exception:
        logging.exception("failed to write security event")

    logging.warning(
        "ICE_REPORT_SECURITY_EVENT type=%s token_hash=%s delivery_id=%s email_hash=%s reason=%s",
        event_type,
        _log_fingerprint(token),
        delivery_id,
        _log_fingerprint(email),
        reason,
    )


def _latest_pin_issue_for_email(token: str, email: str) -> datetime | None:
    db = firestore.Client()
    query = (
        db.collection(_otp_collection_name())
        .where("token", "==", token)
        .where("email", "==", email)
        .limit(20)
    )

    latest = None

    for doc in query.stream():
        data = doc.to_dict() or {}
        created_at = data.get("created_at")
        if not created_at:
            continue

        if latest is None or created_at > latest:
            latest = created_at

    return latest


def _check_pin_resend_interval(token: str, email: str) -> tuple[bool, str]:
    interval_seconds = _int_env("OTP_RESEND_INTERVAL_SECONDS", 60)
    if interval_seconds <= 0:
        return True, ""

    latest = _latest_pin_issue_for_email(token, email)
    if not latest:
        return True, ""

    now = _now_utc()
    elapsed = int((now - latest).total_seconds())
    if elapsed >= interval_seconds:
        return True, ""

    retry_after = interval_seconds - elapsed
    return False, f"PIN再発行は{retry_after}秒後に再試行してください。"


def _rate_limit_collection_name() -> str:
    return os.environ.get("OTP_RATE_LIMIT_COLLECTION", "otp_rate_limits")


def _rate_limit_key(kind: str, value: str) -> str:
    raw = f"{kind}:{value}"
    return _hash_value(raw)


def _check_and_increment_rate_limit(
    *,
    kind: str,
    value: str,
    limit: int,
    window_seconds: int,
) -> tuple[bool, str]:
    if not value or limit <= 0 or window_seconds <= 0:
        return True, ""

    now = _now_utc()
    bucket_start_seconds = int(now.timestamp()) // window_seconds * window_seconds
    bucket_start = datetime.fromtimestamp(bucket_start_seconds, timezone.utc)
    bucket_end = bucket_start + timedelta(seconds=window_seconds)

    doc_id = _rate_limit_key(kind, f"{value}:{bucket_start_seconds}")
    db = firestore.Client()
    ref = db.collection(_rate_limit_collection_name()).document(doc_id)

    @firestore.transactional
    def update_in_transaction(transaction):
        snapshot = ref.get(transaction=transaction)

        if snapshot.exists:
            data = snapshot.to_dict() or {}
            count = int(data.get("count") or 0)

            if count >= limit:
                return False, count

            transaction.update(
                ref,
                {
                    "count": firestore.Increment(1),
                    "updated_at": now,
                },
            )
            return True, count + 1

        transaction.set(
            ref,
            {
                "kind": kind,
                "value_hash": _hash_value(value),
                "count": 1,
                "limit": limit,
                "window_seconds": window_seconds,
                "window_start": bucket_start,
                "window_end": bucket_end,
                "created_at": now,
                "updated_at": now,
            },
        )
        return True, 1

    transaction = db.transaction()
    allowed, count = update_in_transaction(transaction)

    if allowed:
        return True, ""

    retry_after = max(1, int((bucket_end - now).total_seconds()))
    return False, f"PIN発行回数が上限に達しました。{retry_after}秒後に再試行してください。"


def _check_pin_request_rate_limits(email: str) -> tuple[bool, str]:
    ip = _get_client_ip()

    checks = [
        (
            "ip_1m",
            ip,
            _int_env("OTP_RATE_LIMIT_IP_PER_MINUTE", 3),
            60,
        ),
        (
            "ip_10m",
            ip,
            _int_env("OTP_RATE_LIMIT_IP_PER_10_MINUTES", 10),
            600,
        ),
        (
            "email_1m",
            email,
            _int_env("OTP_RATE_LIMIT_EMAIL_PER_MINUTE", 3),
            60,
        ),
        (
            "email_10m",
            email,
            _int_env("OTP_RATE_LIMIT_EMAIL_PER_10_MINUTES", 10),
            600,
        ),
    ]

    for kind, value, limit, window_seconds in checks:
        allowed, message = _check_and_increment_rate_limit(
            kind=kind,
            value=value,
            limit=limit,
            window_seconds=window_seconds,
        )
        if not allowed:
            return False, message

    return True, ""


def _issue_pin(token: str, delivery_id: str, email: str) -> dict:
    now = _now_utc()
    ttl_minutes = _int_env("OTP_PIN_TTL_MINUTES", 10)
    max_attempts = _int_env("OTP_MAX_ATTEMPTS", 5)
    pin = f"{secrets.randbelow(1000000):06d}"

    record = {
        "token": token,
        "delivery_id": delivery_id,
        "email": email,
        "pin_hash": _hash_value(pin),
        "expires_at": now + timedelta(minutes=ttl_minutes),
        "attempt_count": 0,
        "max_attempts": max_attempts,
        "used": False,
        "created_at": now,
        "ip": _get_client_ip(),
        "user_agent": request.headers.get("User-Agent", ""),
    }

    revoked_count = _revoke_existing_challenges(token, email)

    db = firestore.Client()
    ref = db.collection(_otp_collection_name()).document()
    ref.set(record)

    logging.warning(
        "ICE_REPORT_OTP_PIN issued token_hash=%s delivery_id=%s email_hash=%s expires_at=%s revoked_previous=%s",
        _log_fingerprint(token),
        delivery_id,
        _log_fingerprint(email),
        record["expires_at"].isoformat(),
        revoked_count,
    )

    delivery_result = send_otp_pin_email(
        to_email=email,
        pin=pin,
        ttl_minutes=ttl_minutes,
        token=token,
        delivery_id=delivery_id,
    )

    logging.warning(
        "ICE_REPORT_OTP_DELIVERY_SENT token_hash=%s delivery_id=%s email_hash=%s provider=%s provider_message_id=%s",
        _log_fingerprint(token),
        delivery_id,
        _log_fingerprint(email),
        delivery_result.provider,
        delivery_result.provider_message_id,
    )

    return {
        "challenge_id": ref.id,
        "expires_at": record["expires_at"],
        "ttl_minutes": ttl_minutes,
        "provider": delivery_result.provider,
        "provider_message_id": delivery_result.provider_message_id,
    }


def _find_valid_challenge(token: str, email: str) -> tuple[str | None, dict | None, str | None]:
    now = _now_utc()

    db = firestore.Client()
    query = (
        db.collection(_otp_collection_name())
        .where("token", "==", token)
        .where("email", "==", email)
        .where("used", "==", False)
        .limit(20)
    )

    newest_id = None
    newest = None

    for doc in query.stream():
        data = doc.to_dict() or {}
        if data.get("revoked"):
            continue

        expires_at = data.get("expires_at")
        max_attempts = int(data.get("max_attempts") or _int_env("OTP_MAX_ATTEMPTS", 5))
        attempt_count = int(data.get("attempt_count") or 0)

        if expires_at and expires_at < now:
            continue

        if attempt_count >= max_attempts:
            continue

        if newest is None or data.get("created_at") > newest.get("created_at"):
            newest_id = doc.id
            newest = data

    if newest:
        return newest_id, newest, None

    return None, None, "PINが見つからないか、有効期限切れです。もう一度PINを発行してください。"


def _create_download_session(token: str, delivery_id: str, email: str) -> tuple[str, datetime]:
    now = _now_utc()
    ttl_minutes = _int_env("DOWNLOAD_SESSION_TTL_MINUTES", 15)
    session_token = secrets.token_urlsafe(32)
    expires_at = now + timedelta(minutes=ttl_minutes)

    record = {
        "token": token,
        "delivery_id": delivery_id,
        "email": email,
        "session_hash": _hash_value(session_token),
        "expires_at": expires_at,
        "used": False,
        "created_at": now,
        "ip": _get_client_ip(),
        "user_agent": request.headers.get("User-Agent", ""),
    }

    db = firestore.Client()
    db.collection(_download_sessions_collection_name()).document().set(record)

    return session_token, expires_at


def _find_download_session(token: str, session_token: str) -> tuple[str | None, dict | None]:
    if not session_token:
        return None, None

    now = _now_utc()
    session_hash = _hash_value(session_token)

    db = firestore.Client()
    query = (
        db.collection(_download_sessions_collection_name()).where("token", "==", token).where("session_hash", "==", session_hash).limit(5)
    )

    for doc in query.stream():
        data = doc.to_dict() or {}
        expires_at = data.get("expires_at")

        if data.get("used"):
            continue

        if expires_at and expires_at < now:
            continue

        return doc.id, data

    return None, None


@app.get("/d/<token>")
def download_form(token: str):
    delivery_id, delivery = find_delivery_by_token(token)

    if not delivery:
        return _render_otp_page(
            token,
            error="配布URLが見つかりません。",
        ), 404

    return _render_otp_page(token, delivery=delivery)


@app.post("/d/<token>/request-pin")
def request_download_pin(token: str):
    email = _normalize_email(request.form.get("email") or "")

    if not email:
        return _render_otp_page(
            token,
            error="メールアドレスを入力してください。",
        ), 400

    delivery_id, delivery = find_delivery_by_token(token)

    if not delivery:
        _log_security_event(
            event_type="otp_request_denied",
            token=token,
            email=email,
            reason="delivery_not_found",
        )
        return _render_otp_page(
            token,
            email=email,
            error="配布URLが見つかりません。",
        ), 404

    allowed, message = validate_delivery_access(
        delivery,
        email,
    )

    if not allowed:
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            error=message,
        ), 403

    resend_allowed, resend_message = _check_pin_resend_interval(token, email)
    if not resend_allowed:
        _log_security_event(
            event_type="otp_resend_interval_blocked",
            token=token,
            delivery_id=delivery_id,
            email=email,
            reason="resend_interval",
            detail={"message": resend_message},
        )
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            error=resend_message,
        ), 429

    rate_allowed, rate_message = _check_pin_request_rate_limits(email)
    if not rate_allowed:
        _log_security_event(
            event_type="otp_rate_limited",
            token=token,
            delivery_id=delivery_id,
            email=email,
            reason="rate_limit",
            detail={"message": rate_message},
        )
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            error=rate_message,
        ), 429

    try:
        pin_result = _issue_pin(token, delivery_id, email)
    except MailDeliveryError as exc:
        _log_security_event(
            event_type="otp_delivery_failed",
            token=token,
            delivery_id=delivery_id,
            email=email,
            reason=exc.safe_reason,
        )
        logging.exception(
            "failed to deliver otp pin token_hash=%s delivery_id=%s email_hash=%s safe_reason=%s retryable=%s provider_error_code=%s",
            _log_fingerprint(token),
            delivery_id,
            _log_fingerprint(email),
            exc.safe_reason,
            exc.retryable,
            exc.provider_error_code,
        )
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            error="PIN送信に失敗しました。少し待ってから再試行してください。",
        ), 503

    _log_security_event(
        event_type="otp_delivery_sent",
        token=token,
        delivery_id=delivery_id,
        email=email,
        reason=pin_result.get("provider", "unknown"),
        detail={
            "provider": pin_result.get("provider", ""),
            "provider_message_id": pin_result.get("provider_message_id", ""),
        },
    )

    return _render_otp_page(
        token,
        email=email,
        delivery=delivery,
        step="pin",
        message="PINを送信しました。メールをご確認のうえ認証を続けてください。",
    )


@app.post("/d/<token>/verify-pin")
def verify_download_pin(token: str):
    email = _normalize_email(request.form.get("email") or "")
    pin = (request.form.get("pin") or "").strip()

    if not email:
        return _render_otp_page(
            token,
            error="メールアドレスを入力してください。",
        ), 400

    if not pin:
        return _render_otp_page(
            token,
            email=email,
            step="pin",
            error="PINを入力してください。",
        ), 400

    delivery_id, delivery = find_delivery_by_token(token)

    if not delivery:
        return _render_otp_page(
            token,
            email=email,
            step="pin",
            error="配布URLが見つかりません。",
        ), 404

    allowed, message = validate_delivery_access(
        delivery,
        email,
    )

    if not allowed:
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            step="pin",
            error=message,
        ), 403

    challenge_id, challenge, challenge_error = _find_valid_challenge(token, email)

    if not challenge:
        _log_security_event(
            event_type="otp_verify_denied",
            token=token,
            delivery_id=delivery_id,
            email=email,
            reason="challenge_not_found_or_expired",
            detail={"message": challenge_error or "PINが無効です。"},
        )
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            step="pin",
            error=challenge_error or "PINが無効です。",
        ), 403

    db = firestore.Client()
    challenge_ref = db.collection(_otp_collection_name()).document(challenge_id)

    if not hmac.compare_digest(challenge.get("pin_hash") or "", _hash_value(pin)):
        challenge_ref.update({
            "attempt_count": firestore.Increment(1),
            "last_failed_at": _now_utc(),
        })

        max_attempts = int(challenge.get("max_attempts") or _int_env("OTP_MAX_ATTEMPTS", 5))
        current_attempts = int(challenge.get("attempt_count") or 0) + 1

        if current_attempts >= max_attempts:
            error_message = "PINの入力回数が上限に達しました。もう一度PINを発行してください。"
            reason = "max_attempts_reached"
        else:
            remaining = max_attempts - current_attempts
            error_message = f"PINが正しくありません。残り{remaining}回です。"
            reason = "wrong_pin"

        _log_security_event(
            event_type="otp_verify_failed",
            token=token,
            delivery_id=delivery_id,
            email=email,
            reason=reason,
            detail={
                "attempt_count": current_attempts,
                "max_attempts": max_attempts,
            },
        )

        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            step="pin",
            error=error_message,
        ), 403

    challenge_ref.update({
        "used": True,
        "verified_at": _now_utc(),
    })

    _log_security_event(
        event_type="otp_verify_success",
        token=token,
        delivery_id=delivery_id,
        email=email,
        reason="verified",
    )

    session_token, session_expires_at = _create_download_session(
        token,
        delivery_id,
        email,
    )

    response = make_response(redirect(f"/d/{token}/download", code=302))
    response.set_cookie(
        _cookie_name(token),
        session_token,
        max_age=_int_env("DOWNLOAD_SESSION_TTL_MINUTES", 15) * 60,
        expires=session_expires_at,
        secure=True,
        httponly=True,
        samesite="Lax",
    )

    return response


@app.get("/d/<token>/download")
def download_file(token: str):
    cookie_value = request.cookies.get(_cookie_name(token), "")
    session_id, session = _find_download_session(token, cookie_value)

    if not session:
        _log_security_event(
            event_type="download_session_denied",
            token=token,
            reason="session_missing_or_expired",
        )
        return _render_otp_page(
            token,
            error="ダウンロード認証が未完了、またはsessionが期限切れです。もう一度PIN認証してください。",
        ), 403

    delivery_id, delivery = find_delivery_by_token(token)

    if not delivery:
        return _render_otp_page(
            token,
            error="配布URLが見つかりません。",
        ), 404

    email = session.get("email") or ""

    allowed, message = validate_delivery_access(
        delivery,
        email,
    )

    if not allowed:
        return _render_otp_page(
            token,
            email=email,
            delivery=delivery,
            error=message,
        ), 403

    version = get_current_version(delivery)

    signed_url = make_signed_download_url(version)

    log_download(
        delivery_id=delivery_id,
        delivery=delivery,
        version=version,
        email=email,
        request=request,
    )

    _log_security_event(
        event_type="download_session_success",
        token=token,
        delivery_id=delivery_id,
        email=email,
        reason="signed_url_redirect",
        detail={
            "version": version.get("version"),
            "file_name": version.get("file_name"),
        },
    )

    if _bool_env("DOWNLOAD_SESSION_ONE_TIME", True) and session_id:
        db = firestore.Client()
        db.collection(_download_sessions_collection_name()).document(session_id).update({
            "used": True,
            "used_at": _now_utc(),
        })

    return redirect(signed_url, code=302)


@app.post("/d/<token>")
def legacy_download_file(token: str):
    return request_download_pin(token)
