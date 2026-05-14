from __future__ import annotations

import argparse
import json
import math
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from google.cloud import bigquery, storage
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = BASE_DIR / "templates" / "template.xlsx"
DEFAULT_OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", "/tmp/report_output"))
PAID_SQL_PATH = BASE_DIR / "sql" / "paid.sql"
FREE_SQL_PATH = BASE_DIR / "sql" / "free.sql"

PAID_SOURCE_SHEET = "見込み_yymm_PLUS"
FREE_SOURCE_SHEET = "無料_yymm_PLUS"

PAID_TOTAL_COLS = range(5, 11)


def previous_month_base(today: Optional[date] = None) -> date:
    if today is None:
        today = date.today()
    first_day_this_month = today.replace(day=1)
    return first_day_this_month - timedelta(days=1)


def build_names(today: Optional[date] = None, output_filename: Optional[str] = None) -> dict[str, str]:
    if today is None:
        today = date.today()

    target_month = previous_month_base(today)
    yymmdd = today.strftime("%y%m%d")
    yymm = target_month.strftime("%y%m")

    default_output_file = f"ダウンロード数入力シート_{yymmdd}_ICE入力済み_plus.xlsx"

    return {
        "yymmdd": yymmdd,
        "yymm": yymm,
        "output_file": output_filename or default_output_file,
        "paid_sheet": f"見込み_{yymm}_PLUS",
        "free_sheet": f"無料_{yymm}_PLUS",
        "report_month": target_month.strftime("%Y-%m"),
    }


def read_sql(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def run_query(client: bigquery.Client, sql_path: Path, dry_run: bool = False) -> pd.DataFrame:
    sql = read_sql(sql_path)
    if dry_run:
        job_config = bigquery.QueryJobConfig(dry_run=True, use_query_cache=False)
        client.query(sql, job_config=job_config)
        return pd.DataFrame()
    return client.query(sql).to_dataframe()


def get_single_table(ws: Worksheet):
    tables = list(ws.tables.values())
    if len(tables) != 1:
        raise ValueError(f"{ws.title}: テーブル数が1つではありません。現在 {len(tables)}個です。")
    return tables[0]


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    for col in normalized.columns:
        if pd.api.types.is_datetime64_any_dtype(normalized[col]):
            normalized[col] = normalized[col].dt.tz_localize(None)
    normalized = normalized.astype(object)
    return normalized.where(pd.notna(normalized), None)


def to_excel_value(value):
    if value is None or value is pd.NA:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            return value
    return value


def write_paid_total_values(ws: Worksheet, data_start_row: int, data_end_row: int, total_row: int) -> None:
    for col in PAID_TOTAL_COLS:
        total = 0
        for row in range(data_start_row, data_end_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            try:
                total += value
            except TypeError as exc:
                cell = f"{get_column_letter(col)}{row}"
                raise ValueError(f"{ws.title}: 集計対象セル {cell} が数値ではありません: {value!r}") from exc
        ws.cell(row=total_row, column=col).value = total


def paste_df_to_table(ws: Worksheet, df: pd.DataFrame, expected_cols: int, has_total_row: bool) -> None:
    df = normalize_dataframe(df)

    if len(df.columns) != expected_cols:
        raise ValueError(f"{ws.title}: {expected_cols}列必要です。現在 {len(df.columns)}列です。")

    table = get_single_table(ws)
    min_col, header_row, max_col, old_table_end_row = range_boundaries(table.ref)

    actual_table_cols = max_col - min_col + 1
    if actual_table_cols != expected_cols:
        raise ValueError(f"{ws.title}: Excelテーブルは{expected_cols}列必要です。現在 {actual_table_cols}列です。")

    data_start_row = header_row + 1
    old_data_end_row = old_table_end_row - 1 if has_total_row else old_table_end_row
    old_capacity = max(old_data_end_row - data_start_row + 1, 0)
    data_rows = df.values.tolist()
    data_count = len(data_rows)

    if data_count == 0:
        raise ValueError(f"{ws.title}: BigQuery結果が0件です。空のレポート生成を止めました。")

    if data_count > old_capacity:
        insert_count = data_count - old_capacity
        insert_at = old_table_end_row if has_total_row else old_table_end_row + 1
        ws.insert_rows(insert_at, insert_count)
        old_table_end_row += insert_count
        old_data_end_row += insert_count

    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=old_data_end_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None

    for r_idx, row_values in enumerate(data_rows, start=data_start_row):
        for c_idx, value in enumerate(row_values, start=min_col):
            ws.cell(row=r_idx, column=c_idx).value = to_excel_value(value)

    if has_total_row:
        new_table_end_row = data_start_row + data_count
        if new_table_end_row < old_table_end_row:
            ws.delete_rows(new_table_end_row, old_table_end_row - new_table_end_row)
        data_end_row = new_table_end_row - 1
        write_paid_total_values(ws, data_start_row, data_end_row, new_table_end_row)
    else:
        new_table_end_row = data_start_row + data_count - 1
        if new_table_end_row < old_table_end_row:
            ws.delete_rows(new_table_end_row + 1, old_table_end_row - new_table_end_row)

    start_col_letter = get_column_letter(min_col)
    end_col_letter = get_column_letter(max_col)
    table.ref = f"{start_col_letter}{header_row}:{end_col_letter}{new_table_end_row}"

    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def rename_sheet_if_needed(ws: Worksheet, new_name: str) -> None:
    if ws.title != new_name:
        ws.title = new_name


def create_workbook(
    template_path: Path,
    output_path: Path,
    df_paid: pd.DataFrame,
    df_free: pd.DataFrame,
    today: Optional[date] = None,
    output_filename: Optional[str] = None,
) -> dict[str, str | int]:
    names = build_names(today, output_filename=output_filename)

    wb = load_workbook(template_path)

    if PAID_SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"テンプレートにシートがありません: {PAID_SOURCE_SHEET}")

    if FREE_SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"テンプレートにシートがありません: {FREE_SOURCE_SHEET}")

    ws_paid = wb[PAID_SOURCE_SHEET]
    ws_free = wb[FREE_SOURCE_SHEET]

    rename_sheet_if_needed(ws_paid, names["paid_sheet"])
    rename_sheet_if_needed(ws_free, names["free_sheet"])

    paste_df_to_table(ws_paid, df_paid, expected_cols=10, has_total_row=True)
    paste_df_to_table(ws_free, df_free, expected_cols=6, has_total_row=False)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "output_file": names["output_file"],
        "paid_sheet": names["paid_sheet"],
        "free_sheet": names["free_sheet"],
        "report_month": names["report_month"],
        "paid_rows": len(df_paid),
        "free_rows": len(df_free),
    }


def upload_to_gcs(local_path: Path, bucket_name: str, object_name: str) -> str:
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(object_name)
    blob.upload_from_filename(
        str(local_path),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return f"gs://{bucket_name}/{object_name}"


def generate_report(
    project_id: str,
    bucket_name: Optional[str] = None,
    object_prefix: str = "reports/plus",
    template_path: Path = DEFAULT_TEMPLATE,
    today: Optional[date] = None,
    dry_run: bool = False,
    output_filename: Optional[str] = None,
) -> dict:
    if today is None:
        today = date.today()

    names = build_names(today, output_filename=output_filename)
    client = bigquery.Client(project=project_id)

    if dry_run:
        run_query(client, PAID_SQL_PATH, dry_run=True)
        run_query(client, FREE_SQL_PATH, dry_run=True)
        return {"status": "dry_run_completed"}

    df_paid = run_query(client, PAID_SQL_PATH)
    df_free = run_query(client, FREE_SQL_PATH)

    output_path = DEFAULT_OUTPUT_DIR / names["output_file"]

    result = create_workbook(
        template_path,
        output_path,
        df_paid,
        df_free,
        today=today,
        output_filename=output_filename,
    )

    result["local_path"] = str(output_path)

    if bucket_name:
        object_name = f"{object_prefix.rstrip('/')}/{names['yymm']}/{names['output_file']}"
        result["gcs_uri"] = upload_to_gcs(output_path, bucket_name, object_name)
        result["gcs_bucket"] = bucket_name
        result["gcs_object"] = object_name

    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="ジャンプ＋月次ExcelレポートをBigQueryから生成し、必要に応じてGCSへ保存します。")
    parser.add_argument("--project", default=os.environ.get("PROJECT_ID"), help="BigQueryジョブを実行するGCPプロジェクトID")
    parser.add_argument("--bucket", default=os.environ.get("BUCKET_NAME"), help="生成Excelを保存するGCS bucket名")
    parser.add_argument("--object-prefix", default=os.environ.get("OBJECT_PREFIX", "reports/plus"), help="GCS object prefix")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Excelテンプレートのパス")
    parser.add_argument("--today", default=None, help="生成日をYYYY-MM-DDで上書きします。未指定なら実行日")
    parser.add_argument("--dry-run", action="store_true", help="BigQueryのdry runだけ実行します。Excelは生成しません。")
    parser.add_argument("--output-filename", default=None, help="生成するExcelファイル名を指定します。")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if not args.project:
        raise ValueError("--project または環境変数 PROJECT_ID が必要です。")

    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()

    result = generate_report(
        project_id=args.project,
        bucket_name=args.bucket,
        object_prefix=args.object_prefix,
        template_path=Path(args.template),
        today=today,
        dry_run=args.dry_run,
        output_filename=args.output_filename,
    )

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()