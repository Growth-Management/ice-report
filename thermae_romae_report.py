from __future__ import annotations

import copy
import math
import os
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from google.cloud import bigquery
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

THERMAE_REPORT_ID = "thermae-romae"
THERMAE_REPORT_NAME = "テルマエ・ロマエ月次販売報告書"
DEFAULT_THERMAE_TEMPLATE_FILE_ID = "1KvfIA96o17oHfTp5dWMCByL8THxU_Txp"
DEFAULT_THERMAE_OUTPUT_FOLDER_ID = "12kjj_xdQ-O6QAFl5QvDWXn4dUIYGlMCa"
DEFAULT_THERMAE_SOURCE_TABLE = (
    "jumpplus-4a5f4.dataset_datamart_tables.report_plus_monthly_coin_content_report"
)
DEFAULT_THERMAE_WORK_IDS = (100040643, 100040644)

INVOICE_SHEET = "支払通知書"
DETAIL_SHEET = "売上明細"
DETAIL_HEADERS = (
    "売上月/売上日",
    "出版社名",
    "書籍コード",
    "タイトル名",
    "単価（税抜）",
    "売上件数",
    "支払額（税抜）",
)
DETAIL_TOTAL_LABELS = ("支払額計", "消費税額（※支払額計×0.1）", "税込計")
DETAIL_DATA_START_ROW = 2
DETAIL_TOTAL_START_ROW = 56
DETAIL_TOTAL_LABEL_COL = 6
DETAIL_TOTAL_AMOUNT_COL = 7
INVOICE_GENERATED_DATE_CELL = "G3"
INVOICE_PAYMENT_DUE_CELL = "B53"
THERMAE_FIXED_DETAIL_ITEMS: tuple[tuple[str, str], ...] = (
    ("04726127A01564800000", "テルマエ・ロマエ 1"),
    ("04726770A01564900000", "テルマエ・ロマエ 2"),
    ("04727232A01565000000", "テルマエ・ロマエ 3"),
    ("04727515A01541000000", "テルマエ・ロマエ 4"),
    ("04728225A01487700000", "テルマエ・ロマエ 5"),
    ("04728895A02356000000", "テルマエ・ロマエ 6"),
    ("04726127A01564801520", "[第1話-①]テルマエ・ロマエ"),
    ("04726127A01564802520", "[第1話-②]テルマエ・ロマエ"),
    ("04726127A01564803520", "[第2話-①]テルマエ・ロマエ"),
    ("04726127A01564804520", "[第2話-②]テルマエ・ロマエ"),
    ("04726127A01564805520", "[第3話-①]テルマエ・ロマエ"),
    ("04726127A01564806520", "[第3話-②]テルマエ・ロマエ"),
    ("04726127A01564807520", "[第4話-①]テルマエ・ロマエ"),
    ("04726127A01564808520", "[第4話-②]テルマエ・ロマエ"),
    ("04726127A01564809520", "[第5話-①]テルマエ・ロマエ"),
    ("04726127A01564810520", "[第5話-②]テルマエ・ロマエ"),
    ("04726770A01564901520", "[第6話-①]テルマエ・ロマエ"),
    ("04726770A01564902520", "[第6話-②]テルマエ・ロマエ"),
    ("04726770A01564903520", "[第7話-①]テルマエ・ロマエ"),
    ("04726770A01564904520", "[第7話-②]テルマエ・ロマエ"),
    ("04726770A01564905520", "[第8話-①]テルマエ・ロマエ"),
    ("04726770A01564906520", "[第8話-②]テルマエ・ロマエ"),
    ("04726770A01564907520", "[第9話-①]テルマエ・ロマエ"),
    ("04726770A01564908520", "[第9話-②]テルマエ・ロマエ"),
    ("04726770A01564909520", "[第10話-①]テルマエ・ロマエ"),
    ("04726770A01564910520", "[第10話-②]テルマエ・ロマエ"),
    ("04727232A01565001520", "[第11話]テルマエ・ロマエ"),
    ("04727232A01565002520", "[第12話]テルマエ・ロマエ"),
    ("04727232A01565003520", "[第13話]テルマエ・ロマエ"),
    ("04727232A01565004520", "[第14話]テルマエ・ロマエ"),
    ("04727232A01565005520", "[第15話]テルマエ・ロマエ"),
    ("04727232A01565006520", "[第16話]テルマエ・ロマエ"),
    ("04727232A01565007520", "[第17話]テルマエ・ロマエ"),
    ("04727515A01541001520", "[第18話]テルマエ・ロマエ"),
    ("04727515A01541002520", "[第19話]テルマエ・ロマエ"),
    ("04727515A01541003520", "[第20話]テルマエ・ロマエ"),
    ("04727515A01541004520", "[第21話]テルマエ・ロマエ"),
    ("04727515A01541005520", "[第22話]テルマエ・ロマエ"),
    ("04727515A01541006520", "[第23話]テルマエ・ロマエ"),
    ("04727515A01541007520", "[第24話]テルマエ・ロマエ"),
    ("04728225A01487701520", "[第25話]テルマエ・ロマエ"),
    ("04728225A01487702520", "[第26話]テルマエ・ロマエ"),
    ("04728225A01487703520", "[第27話]テルマエ・ロマエ"),
    ("04728225A01487704520", "[第28話]テルマエ・ロマエ"),
    ("04728225A01487705520", "[第29話]テルマエ・ロマエ"),
    ("04728225A01487706520", "[第30話]テルマエ・ロマエ"),
    ("04728225A01487707520", "[第31話]テルマエ・ロマエ"),
    ("04728895A02356001520", "[第32話]テルマエ・ロマエ"),
    ("04728895A02356002520", "[第33話]テルマエ・ロマエ"),
    ("04728895A02356003520", "[第34話]テルマエ・ロマエ"),
    ("04728895A02356004520", "[第35話]テルマエ・ロマエ"),
    ("04728895A02356005520", "[第36話]テルマエ・ロマエ"),
    ("04728895A02356006520", "[第37話]テルマエ・ロマエ"),
    ("04728895A02356007520", "[第38話]テルマエ・ロマエ"),
)

THERMAE_SQL = """
with base as (
    select
        purchase_date_month_jst
        , work_id
        , work_title
        , name as title_name
        , unit_price
        , coalesce(pay_coins_total, 0) as pay_coins_total
        , coalesce(pay_bonus_coins_total, 0) as pay_bonus_coins_total
        , coalesce(free_bonus_coins_total, 0) as free_bonus_coins_total
    from
        `{source_table}`
    where 1=1
        and purchase_date_month_jst = @target_month
        and cast(work_id as int64) in unnest(@work_ids)
)

, aggregated as (
    select
        purchase_date_month_jst as report_month
        , format_date('%Y年%-m月', purchase_date_month_jst) as sales_month_label
        , 'KADOKAWA' as publisher_name
        , work_id
        , work_title
        , title_name
        , unit_price as unit_price_tax_included
        , cast(round(unit_price / 1.1) as int64) as unit_price_tax_excluded
        , sum(
            pay_coins_total
            + pay_bonus_coins_total
            + free_bonus_coins_total
        ) as target_coins
        , safe_divide(
            sum(
                pay_coins_total
                + pay_bonus_coins_total
                + free_bonus_coins_total
            ),
            unit_price
        ) as sales_count_raw
    from
        base
    group by all
)

select
    report_month
    , sales_month_label
    , publisher_name
    , work_id
    , work_title
    , title_name
    , unit_price_tax_included
    , unit_price_tax_excluded
    , target_coins
    , cast(round(sales_count_raw) as int64) as sales_count
    , cast(
        round(
            cast(round(sales_count_raw) as int64)
            * unit_price_tax_excluded
            * 0.55
        ) as int64
    ) as payment_amount_tax_excluded
from
    aggregated
order by
    case
        when work_id = 100040644 then 1
        when work_id = 100040643 then 2
        else 9
    end
    , title_name
"""


class ThermaeReportError(Exception):
    def __init__(self, code: str, message: str | None = None, *, status_code: int = 400, **details: Any) -> None:
        super().__init__(message or code)
        self.code = code
        self.status_code = status_code
        self.details = details


def previous_month_first(today: date | None = None) -> date:
    today = today or date.today()
    first_this_month = today.replace(day=1)
    previous_month_last = first_this_month - timedelta(days=1)
    return previous_month_last.replace(day=1)


def parse_target_month(value: str | None = None, *, today: date | None = None) -> date:
    if not value:
        return previous_month_first(today)
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ThermaeReportError("invalid_target_month", "target_month must be YYYY-MM-DD") from exc
    if parsed.day != 1:
        raise ThermaeReportError("invalid_target_month", "target_month must be the first day of month")
    return parsed


def _add_months(value: date, months: int) -> date:
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def month_end(value: date) -> date:
    return _add_months(value, 1) - timedelta(days=1)


def payment_due_month_end(target_month: date) -> date:
    return month_end(_add_months(target_month, 3))


def period_label(target_month: date) -> str:
    end = month_end(target_month)
    return f"{target_month:%Y年%m月%d日}〜{end:%m月%d日}"


def payment_due_text(generated_date: date) -> str:
    due = payment_due_month_end(generated_date)
    return f"※御支払いは{due.year}年{due.month}月末を予定しております。"


def output_file_name(target_month: date) -> str:
    return f"KADOKAWA様_少年ジャンプ＋「テルマエ・ロマエ」販売報告書_{target_month.year}年{target_month.month}月分.xlsx"


def parse_work_ids(value: str | None = None) -> list[int]:
    if not value:
        return list(DEFAULT_THERMAE_WORK_IDS)
    items = []
    for raw in value.split(","):
        raw = raw.strip()
        if raw:
            items.append(int(raw))
    return items


def source_table() -> str:
    return os.environ.get("THERMAE_SOURCE_TABLE", DEFAULT_THERMAE_SOURCE_TABLE)


def run_thermae_query(
    *,
    project_id: str,
    target_month: date,
    table: str | None = None,
    work_ids: list[int] | None = None,
) -> list[dict[str, Any]]:
    client = bigquery.Client(project=project_id)
    query = THERMAE_SQL.format(source_table=table or source_table())
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("target_month", "DATE", target_month),
            bigquery.ArrayQueryParameter("work_ids", "INT64", work_ids or parse_work_ids(os.environ.get("THERMAE_WORK_IDS"))),
        ]
    )
    frame = client.query(query, job_config=job_config).to_dataframe()
    records = frame.to_dict("records")
    if not records:
        raise ThermaeReportError("no_rows", "no rows returned for target_month")
    return records


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def find_header_row(ws: Worksheet, required_headers: Iterable[str] = DETAIL_HEADERS) -> tuple[int, dict[str, int]]:
    required = tuple(required_headers)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        values = {_cell_text(cell.value): cell.column for cell in row if _cell_text(cell.value)}
        if all(header in values for header in required):
            return row[0].row, values
    raise ThermaeReportError("detail_header_not_found", "売上明細 header row not found")


def extract_book_code_mapping(ws: Worksheet) -> dict[str, str]:
    header_row, columns = find_header_row(ws, ("書籍コード", "タイトル名"))
    code_col = columns["書籍コード"]
    title_col = columns["タイトル名"]
    mapping: dict[str, str] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        title = _cell_text(ws.cell(row=row_idx, column=title_col).value)
        code = _cell_text(ws.cell(row=row_idx, column=code_col).value)
        if title and code and title not in DETAIL_TOTAL_LABELS:
            mapping[title] = code
    return mapping


def sales_month_label(target_month: date) -> str:
    return f"{target_month.year}年{target_month.month}月"


def build_detail_rows(
    records: list[dict[str, Any]],
    book_code_mapping: dict[str, str] | None = None,
    *,
    target_month: date | None = None,
    fixed_items: Iterable[tuple[str, str]] = THERMAE_FIXED_DETAIL_ITEMS,
) -> list[dict[str, Any]]:
    fixed = tuple(fixed_items)
    fixed_titles = {title for _, title in fixed}
    by_title = {_cell_text(record.get("title_name")): record for record in records}
    unexpected_titles = sorted(title for title in by_title if title and title not in fixed_titles)
    if unexpected_titles:
        raise ThermaeReportError(
            "unexpected_title_name",
            "title_name is not in fixed Thermae Romae detail list",
            title_name=unexpected_titles[0],
        )

    default_sales_month = sales_month_label(target_month) if target_month else (
        _cell_text(records[0].get("sales_month_label")) if records else ""
    )
    rows = []
    for book_code, title_name in fixed:
        record = by_title.get(title_name) or {}
        rows.append(
            {
                "売上月/売上日": record.get("sales_month_label") or default_sales_month,
                "出版社名": record.get("publisher_name") or "KADOKAWA",
                "書籍コード": book_code,
                "タイトル名": title_name,
                "単価（税抜）": int(record.get("unit_price_tax_excluded") or 0),
                "売上件数": int(record.get("sales_count") or 0),
                "支払額（税抜）": int(record.get("payment_amount_tax_excluded") or 0),
            }
        )
    return rows


def summarize_detail_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    payment_total = sum(int(row.get("支払額（税抜）") or 0) for row in rows)
    tax = int(round(payment_total * 0.1))
    return {
        "payment_total": payment_total,
        "tax": tax,
        "total_with_tax": payment_total + tax,
        "detail_row_count": len(rows),
    }


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
    for col in range(min_col, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)
        target.border = copy.copy(source.border)
        target.fill = copy.copy(source.fill)
        target.font = copy.copy(source.font)
        target.protection = copy.copy(source.protection)


def _first_total_row(ws: Worksheet, header_row: int) -> int | None:
    labels = set(DETAIL_TOTAL_LABELS)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        for cell in ws[row_idx]:
            if _cell_text(cell.value) in labels:
                return row_idx
    return None


def _border_with_outline(
    border: Border,
    *,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(
        left=left or border.left,
        right=right or border.right,
        top=top or border.top,
        bottom=bottom or border.bottom,
        diagonal=border.diagonal,
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=border.vertical,
        horizontal=border.horizontal,
        start=border.start,
        end=border.end,
    )


def _apply_thick_outline(
    ws: Worksheet,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    thick = Side(style="thick", color="000000")
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _border_with_outline(
                cell.border,
                left=thick if col_idx == min_col else None,
                right=thick if col_idx == max_col else None,
                top=thick if row_idx == min_row else None,
                bottom=thick if row_idx == max_row else None,
            )


def write_detail_sheet(ws: Worksheet, rows: list[dict[str, Any]], summary: dict[str, int]) -> None:
    header_row, columns = find_header_row(ws)
    min_col = min(columns[header] for header in DETAIL_HEADERS)
    max_col = max(columns[header] for header in DETAIL_HEADERS)
    data_start = DETAIL_DATA_START_ROW
    data_capacity = DETAIL_TOTAL_START_ROW - DETAIL_DATA_START_ROW
    if len(rows) > data_capacity:
        raise ThermaeReportError("too_many_detail_rows", "fixed detail rows exceed template capacity")

    clear_max_col = max(max_col, DETAIL_TOTAL_AMOUNT_COL)
    clear_until = DETAIL_TOTAL_START_ROW + len(DETAIL_TOTAL_LABELS) - 1
    for row_idx in range(data_start, clear_until + 1):
        for col in range(min_col, clear_max_col + 1):
            ws.cell(row=row_idx, column=col).value = None

    for row_offset, row_values in enumerate(rows):
        row_idx = data_start + row_offset
        for header in DETAIL_HEADERS:
            ws.cell(row=row_idx, column=columns[header]).value = row_values.get(header)

    totals = (
        ("支払額計", summary["payment_total"]),
        ("消費税額（※支払額計×0.1）", summary["tax"]),
        ("税込計", summary["total_with_tax"]),
    )
    for offset, (label, value) in enumerate(totals):
        row_idx = DETAIL_TOTAL_START_ROW + offset
        ws.cell(row=row_idx, column=DETAIL_TOTAL_LABEL_COL).value = label
        ws.cell(row=row_idx, column=DETAIL_TOTAL_AMOUNT_COL).value = value

    _apply_thick_outline(
        ws,
        min_row=DETAIL_TOTAL_START_ROW,
        max_row=DETAIL_TOTAL_START_ROW + len(DETAIL_TOTAL_LABELS) - 1,
        min_col=DETAIL_TOTAL_LABEL_COL,
        max_col=DETAIL_TOTAL_AMOUNT_COL,
    )


def write_invoice_sheet(
    ws: Worksheet,
    target_month: date,
    summary: dict[str, int],
    *,
    generated_date: date,
) -> None:
    ws[INVOICE_GENERATED_DATE_CELL] = generated_date
    ws["D30"] = period_label(target_month)
    ws["E42"] = summary["payment_total"]
    ws["E43"] = summary["tax"]
    ws["E44"] = summary["total_with_tax"]
    ws[INVOICE_PAYMENT_DUE_CELL] = payment_due_text(generated_date)
    ws.print_area = "A3:G61"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.scale = 92
    ws.print_options.horizontalCentered = True


def create_thermae_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    target_month: date,
    generated_date: date | None = None,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    generated_date = generated_date or date.today()
    workbook = load_workbook(template_path)
    if INVOICE_SHEET not in workbook.sheetnames:
        raise ThermaeReportError("invoice_sheet_not_found", "支払通知書 sheet not found")
    if DETAIL_SHEET not in workbook.sheetnames:
        raise ThermaeReportError("detail_sheet_not_found", "売上明細 sheet not found")

    detail_ws = workbook[DETAIL_SHEET]
    detail_rows = build_detail_rows(records, target_month=target_month)
    summary = summarize_detail_rows(detail_rows)
    write_detail_sheet(detail_ws, detail_rows, summary)
    write_invoice_sheet(workbook[INVOICE_SHEET], target_month, summary, generated_date=generated_date)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)

    return {
        "target_month": target_month.isoformat(),
        "generated_date": generated_date.isoformat(),
        "file_name": output.name,
        "detail_row_count": summary["detail_row_count"],
        "payment_total": summary["payment_total"],
        "tax": summary["tax"],
        "total_with_tax": summary["total_with_tax"],
        "local_path": str(output),
    }


def generate_thermae_romae_report(
    *,
    project_id: str,
    target_month_text: str | None = None,
    today: date | None = None,
    template_file_id: str | None = None,
    output_folder_id: str | None = None,
) -> dict[str, Any]:
    if not project_id:
        raise ThermaeReportError("project_required", "BIGQUERY_PROJECT_ID or PROJECT_ID is required")

    generated_date = today or date.today()
    target_month = parse_target_month(target_month_text, today=generated_date)
    template_id = template_file_id or os.environ.get("THERMAE_TEMPLATE_FILE_ID", DEFAULT_THERMAE_TEMPLATE_FILE_ID)
    folder_id = output_folder_id or os.environ.get("THERMAE_OUTPUT_FOLDER_ID", DEFAULT_THERMAE_OUTPUT_FOLDER_ID)
    file_name = output_file_name(target_month)

    records = run_thermae_query(project_id=project_id, target_month=target_month)

    from drive_io import download_drive_file, upload_xlsx_to_drive

    with tempfile.TemporaryDirectory(prefix="thermae-report-") as tmp_dir:
        tmp_root = Path(tmp_dir)
        template_path = tmp_root / "template.xlsx"
        output_path = tmp_root / file_name
        download_drive_file(template_id, template_path)
        result = create_thermae_workbook(
            template_path=template_path,
            output_path=output_path,
            target_month=target_month,
            generated_date=generated_date,
            records=records,
        )
        uploaded = upload_xlsx_to_drive(
            output_path,
            folder_id=folder_id,
            file_name=file_name,
        )

    return {
        "status": "ok",
        "report": THERMAE_REPORT_ID,
        "report_name": THERMAE_REPORT_NAME,
        "target_month": target_month.isoformat(),
        "generated_date": result["generated_date"],
        "file_id": uploaded.get("id", ""),
        "file_name": uploaded.get("name") or file_name,
        "webViewLink": uploaded.get("webViewLink", ""),
        "detail_row_count": result["detail_row_count"],
        "payment_total": result["payment_total"],
        "tax": result["tax"],
        "total_with_tax": result["total_with_tax"],
    }
